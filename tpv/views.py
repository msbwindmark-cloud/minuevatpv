import json
from decimal import Decimal
from datetime import datetime, date, time, timedelta
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import admin
from django.db import transaction
from django.db.models import Sum, Count, F, Q, Avg
from django.utils import timezone
from django.conf import settings
from .models import (
    CategoriaProducto, Articulo, OperacionVenta, LineaVenta,
    InsumoMateriaPrima, RegistroGasto, Mesa, AuditLog,
    TurnoCaja, PedidoCocina, MenuDelDia, Reserva,
    CuponDescuento, PerfilEmpleado, MetaDiaria, Cliente,
    ColaNumero, PedidoMovil, LineaPedidoMovil, PlanoRestaurante, MesaPosicion,
    MensajeChat, PedidoProveedor, FirmaDigital, PasoReceta, NotificacionPush,
    DescuentoInteligente, ReviewRestaurante, RatingProducto, Combo, ComboItem
)
from .signals import get_ip


def es_empleado(user):
    return user.is_authenticated and (user.is_superuser or PerfilEmpleado.objects.filter(user=user).exists())

def es_gerente(user):
    return user.is_superuser

def es_admin_panel(user):
    return user.is_superuser or (user.is_authenticated and user.is_staff)

def es_cocina(user):
    if user.is_superuser:
        return True
    try:
        perfil = PerfilEmpleado.objects.get(user=user)
        return perfil.rol == 'COCINA'
    except PerfilEmpleado.DoesNotExist:
        return False


@login_required
def vista_terminal_pos(request):
    categorias = CategoriaProducto.objects.all()
    articulos = Articulo.objects.filter(activo=True)
    mesas = Mesa.objects.filter(activa=True).order_by('numero')
    top_vendidos = (
        LineaVenta.objects
        .values('articulo__id', 'articulo__nombre')
        .annotate(total=Sum('unidades'))
        .order_by('-total')[:6]
    )
    top_ids = [t['articulo__id'] for t in top_vendidos]
    return render(request, 'pos_tactil.html', {
        'categorias': categorias,
        'articulos': articulos,
        'mesas': mesas,
        'top_ids': top_ids,
    })


@login_required
@transaction.atomic
def api_registrar_cobro(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        items_pedido = datos.get('items', [])
        forma_pago = datos.get('forma_pago', 'EFECTIVO')
        mesa_id = datos.get('mesa_id', None)
        cupon_codigo = datos.get('cupon', '').strip().upper()

        if not items_pedido:
            return JsonResponse({"error": "El pedido esta vacio"}, status=400)

        mesa_obj = None
        if mesa_id:
            try:
                mesa_obj = Mesa.objects.get(id=mesa_id)
            except Mesa.DoesNotExist:
                pass

        venta = OperacionVenta.objects.create(empleado_caja=request.user, forma_pago=forma_pago, mesa=mesa_obj)
        acumulado_base = Decimal('0.00')
        acumulado_iva = Decimal('0.00')
        acumulado_total = Decimal('0.00')
        alertas_reposicion = []

        for item in items_pedido:
            articulo = Articulo.objects.get(id=item['id'])
            cantidad = int(item['cantidad'])

            precio_final = articulo.precio_con_iva * cantidad
            porcentaje_iva = Decimal(articulo.tipo_iva) / Decimal('100')
            base_calculada = precio_final / (1 + porcentaje_iva)

            acumulado_base += base_calculada
            acumulado_iva += (precio_final - base_calculada)
            acumulado_total += precio_final

            LineaVenta.objects.create(
                venta=venta,
                articulo=articulo,
                unidades=cantidad,
                precio_aplicado_con_iva=articulo.precio_con_iva
            )

            for receta in articulo.receta.all():
                insumo = receta.insumo
                insumo.cantidad_actual -= (receta.cantidad_consumida * cantidad)
                insumo.save()

                if insumo.necesita_reposicion:
                    alertas_reposicion.append({
                        "ingrediente": insumo.nombre,
                        "restante": float(insumo.cantidad_actual),
                        "unidad": insumo.unidad_medida
                    })

        venta.subtotal_base = acumulado_base.quantize(Decimal('0.01'))
        venta.total_impuestos = acumulado_iva.quantize(Decimal('0.01'))

        descuento = Decimal('0.00')
        cupon_obj = None
        if cupon_codigo:
            cupon_obj = CuponDescuento.objects.filter(codigo=cupon_codigo).first()
            if cupon_obj and cupon_obj.esta_disponible:
                descuento = cupon_obj.calcular_descuento(acumulado_total)
                cupon_obj.usos_realizados += 1
                cupon_obj.save()

        venta.total_facturado = (acumulado_total - descuento).quantize(Decimal('0.01'))
        venta.descuento_aplicado = descuento
        venta.cupon = cupon_obj
        venta.hash_seguridad = f"HASH_SECURE_V_{venta.id}_GEN2026"

        cliente_id = datos.get('cliente_id')
        puntos_ganados = 0
        if cliente_id:
            try:
                cliente = Cliente.objects.get(id=cliente_id)
                puntos_ganados = cliente.agregar_puntos(venta.total_facturado)
                venta.cliente = cliente
                venta.puntos_ganados = puntos_ganados
            except Cliente.DoesNotExist:
                pass

        venta.save()

        articulos_detalle = []
        for lv in venta.lineas.select_related('articulo'):
            articulos_detalle.append(f"{lv.unidades}x {lv.articulo.nombre}")

        AuditLog.objects.create(
            usuario=request.user, accion='VENTA', modelo='Venta',
            objeto_id=venta.id,
            descripcion=f"Ticket #{venta.id} cobrado: {venta.total_facturado}€ ({forma_pago})",
            detalles_json={
                'ticket_id': venta.id, 'total': float(venta.total_facturado),
                'forma_pago': forma_pago,
                'mesa': mesa_obj.numero if mesa_obj else None,
                'articulos': articulos_detalle,
            },
            ip_address=get_ip(request),
        )

        return JsonResponse({
            "status": "success",
            "total": float(venta.total_facturado),
            "ticket_id": venta.id,
            "descuento": float(descuento),
            "puntos_ganados": puntos_ganados,
            "alertas": alertas_reposicion
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@user_passes_test(lambda u: u.is_superuser)
def vista_dashboard_gerencia(request):
    from datetime import date

    hoy = date.today()

    ventas_hoy = OperacionVenta.objects.filter(fecha_registro__date=hoy)
    ingresos_hoy = ventas_hoy.aggregate(total=Sum('total_facturado'))['total'] or Decimal('0.00')
    tickets_hoy = ventas_hoy.count()
    ticket_medio = (ingresos_hoy / tickets_hoy) if tickets_hoy > 0 else Decimal('0.00')

    ingresos_total = OperacionVenta.objects.aggregate(total=Sum('total_facturado'))['total'] or Decimal('0.00')
    gastos_total = RegistroGasto.objects.aggregate(total=Sum('importe_total'))['total'] or Decimal('0.00')
    balance = ingresos_total - gastos_total
    total_tickets = OperacionVenta.objects.count()

    iva_efectivo = ventas_hoy.filter(forma_pago='EFECTIVO').aggregate(total=Sum('total_impuestos'))['total'] or Decimal('0.00')
    iva_tarjeta = ventas_hoy.filter(forma_pago='TARJETA').aggregate(total=Sum('total_impuestos'))['total'] or Decimal('0.00')

    efectivo_hoy = ventas_hoy.filter(forma_pago='EFECTIVO').aggregate(total=Sum('total_facturado'))['total'] or Decimal('0.00')
    tarjeta_hoy = ventas_hoy.filter(forma_pago='TARJETA').aggregate(total=Sum('total_facturado'))['total'] or Decimal('0.00')

    ultimas_ventas = OperacionVenta.objects.select_related('empleado_caja').order_by('-fecha_registro')[:15]

    insumos_criticos = InsumoMateriaPrima.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    )

    top_articulos = (
        LineaVenta.objects
        .values('articulo__nombre')
        .annotate(total_vendido=Sum('unidades'))
        .order_by('-total_vendido')[:8]
    )

    ventas_por_dia = (
        OperacionVenta.objects
        .filter(fecha_registro__date__gte=hoy.__class__(hoy.year, hoy.month, max(1, hoy.day - 6)))
        .values('fecha_registro__date')
        .annotate(total=Sum('total_facturado'), tickets=Count('id'))
        .order_by('fecha_registro__date')
    )

    ventas_por_hora = []
    for h in range(8, 21):
        count = ventas_hoy.filter(fecha_registro__hour=h).aggregate(t=Sum('total_facturado'))['t'] or 0
        ventas_por_hora.append({'hora': f'{h}:00', 'total': float(count)})

    top_labels = [a['articulo__nombre'] for a in top_articulos]
    top_values = [int(a['total_vendido']) for a in top_articulos]
    dias_labels = [str(v['fecha_registro__date'].strftime('%d/%m')) for v in ventas_por_dia]
    dias_values = [float(v['total']) for v in ventas_por_dia]

    return render(request, 'dashboard.html', {
        'ingresos_hoy': ingresos_hoy,
        'tickets_hoy': tickets_hoy,
        'ticket_medio': ticket_medio,
        'ingresos_total': ingresos_total,
        'gastos_total': gastos_total,
        'balance': balance,
        'total_tickets': total_tickets,
        'iva_efectivo': iva_efectivo,
        'iva_tarjeta': iva_tarjeta,
        'ultimas_ventas': ultimas_ventas,
        'insumos_criticos': insumos_criticos,
        'top_articulos': top_articulos,
        'ventas_por_dia': list(ventas_por_dia),
        'dias_labels': json.dumps(dias_labels),
        'dias_values': json.dumps(dias_values),
        'top_labels': json.dumps(top_labels),
        'top_values': json.dumps(top_values),
        'ventas_por_hora': json.dumps(ventas_por_hora),
        'efectivo_hoy': float(efectivo_hoy),
        'tarjeta_hoy': float(tarjeta_hoy),
        'hoy': hoy,
    })


def vista_logout(request):
    logout(request)
    return redirect('login')


@login_required
def api_crear_mesa(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        numero = int(datos.get('numero', 0))
        sillas = int(datos.get('sillas', 4))
        if numero < 1:
            return JsonResponse({"error": "Numero invalido"}, status=400)
        mesa, created = Mesa.objects.get_or_create(
            numero=numero, defaults={'sillas': sillas, 'activa': True}
        )
        if not created:
            return JsonResponse({"error": "Esa mesa ya existe"}, status=400)
        return JsonResponse({"status": "success", "id": mesa.id, "numero": mesa.numero, "sillas": mesa.sillas})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@user_passes_test(lambda u: u.is_superuser)
def vista_bitacora(request):
    from django.core.paginator import Paginator

    logs = AuditLog.objects.select_related('usuario').all()

    filtro_accion = request.GET.get('accion', '')
    filtro_modelo = request.GET.get('modelo', '')
    filtro_usuario = request.GET.get('usuario', '')
    buscar = request.GET.get('q', '')

    if filtro_accion:
        logs = logs.filter(accion=filtro_accion)
    if filtro_modelo:
        logs = logs.filter(modelo=filtro_modelo)
    if filtro_usuario:
        logs = logs.filter(usuario__username=filtro_usuario)
    if buscar:
        logs = logs.filter(descripcion__icontains=buscar)

    paginator = Paginator(logs, 50)
    page = paginator.get_page(request.GET.get('page', 1))

    accion_fija = AuditLog.objects.values_list('accion', flat=True).distinct()
    modelo_fija = AuditLog.objects.values_list('modelo', flat=True).distinct()
    usuarios_fijos = AuditLog.objects.values_list('usuario__username', flat=True).distinct()

    return render(request, 'bitacora.html', {
        'page': page,
        'accion_fija': accion_fija,
        'modelo_fija': modelo_fija,
        'usuarios_fijos': usuarios_fijos,
        'filtro_accion': filtro_accion,
        'filtro_modelo': filtro_modelo,
        'filtro_usuario': filtro_usuario,
        'buscar': buscar,
    })


@login_required
def api_abrir_turno(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    turno_abierto = TurnoCaja.objects.filter(cajero=request.user, cerrado=False).first()
    if turno_abierto:
        return JsonResponse({"error": "Ya tienes un turno abierto"}, status=400)
    try:
        datos = json.loads(request.body)
        saldo = Decimal(str(datos.get('saldo_inicial', 0)))
        turno = TurnoCaja.objects.create(cajero=request.user, saldo_inicial=saldo)
        AuditLog.objects.create(
            usuario=request.user, accion='CREAR', modelo='TurnoCaja',
            objeto_id=turno.id,
            descripcion=f"Turno #{turno.id} abierto con {saldo}€",
            detalles_json={'saldo_inicial': float(saldo)},
            ip_address=get_ip(request),
        )
        return JsonResponse({"status": "success", "turno_id": turno.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def vista_cierre_caja(request):
    if request.user.is_superuser:
        turno = TurnoCaja.objects.filter(cerrado=False).order_by('-fecha_apertura').first()
        ver_todos = True
    else:
        turno = TurnoCaja.objects.filter(cajero=request.user, cerrado=False).first()
        ver_todos = False

    ventas_turno = []
    gastos_turno = []
    total_ef = Decimal('0.00')
    total_tj = Decimal('0.00')
    total_gastos = Decimal('0.00')
    total_teorico = Decimal('0.00')
    turno_usuario = turno.cajero if turno else None

    if turno:
        ventas_turno = OperacionVenta.objects.filter(
            empleado_caja=turno.cajero,
            fecha_registro__gte=turno.fecha_apertura
        ).order_by('-fecha_registro')
        gastos_turno = RegistroGasto.objects.filter(
            empleado_autoriza=turno.cajero,
            fecha_gasto__gte=turno.fecha_apertura
        ).order_by('-fecha_gasto')
        total_ef = ventas_turno.filter(forma_pago='EFECTIVO').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        total_tj = ventas_turno.filter(forma_pago='TARJETA').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        total_gastos = gastos_turno.aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
        total_teorico = turno.saldo_inicial + total_ef + total_tj - total_gastos

    historial = TurnoCaja.objects.filter(cerrado=True)
    if not request.user.is_superuser:
        historial = historial.filter(cajero=request.user)
    historial = historial.order_by('-fecha_cierre')[:20]

    return render(request, 'cierre_caja.html', {
        'turno': turno,
        'ventas_turno': ventas_turno,
        'gastos_turno': gastos_turno,
        'total_ef': total_ef,
        'total_tj': total_tj,
        'total_gastos': total_gastos,
        'total_teorico': total_teorico,
        'turno_usuario': turno_usuario,
        'ver_todos': ver_todos,
        'historial': historial,
    })


@login_required
def api_cerrar_turno(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    if request.user.is_superuser:
        turno = TurnoCaja.objects.filter(cerrado=False).order_by('-fecha_apertura').first()
    else:
        turno = TurnoCaja.objects.filter(cajero=request.user, cerrado=False).first()
    if not turno:
        return JsonResponse({"error": "No hay turno abierto"}, status=400)
    try:
        datos = json.loads(request.body)
        total_real = Decimal(str(datos.get('total_real', 0)))
        notas = datos.get('notas', '')

        ventas_turno = OperacionVenta.objects.filter(
            empleado_caja=request.user, fecha_registro__gte=turno.fecha_apertura
        )
        gastos_turno = RegistroGasto.objects.filter(
            empleado_autoriza=request.user, fecha_gasto__gte=turno.fecha_apertura
        )
        total_ef = ventas_turno.filter(forma_pago='EFECTIVO').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        total_tj = ventas_turno.filter(forma_pago='TARJETA').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        total_gastos = gastos_turno.aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
        total_teorico = turno.saldo_inicial + total_ef + total_tj - total_gastos

        turno.fecha_cierre = timezone.now()
        turno.total_efectivo = total_ef
        turno.total_tarjeta = total_tj
        turno.total_gastos = total_gastos
        turno.total_teorico = total_teorico
        turno.total_real = total_real
        turno.tickets_cerrados = ventas_turno.count()
        turno.cerrado = True
        turno.notas = notas
        turno.save()

        AuditLog.objects.create(
            usuario=request.user, accion='EDITAR', modelo='TurnoCaja',
            objeto_id=turno.id,
            descripcion=f"Turno #{turno.id} cerrado | Teorico: {total_teorico}€ | Real: {total_real}€",
            detalles_json={
                'total_efectivo': float(total_ef), 'total_tarjeta': float(total_tj),
                'total_gastos': float(total_gastos), 'total_teorico': float(total_teorico),
                'total_real': float(total_real), 'tickets': ventas_turno.count(),
            },
            ip_address=get_ip(request),
        )
        return JsonResponse({"status": "success", "turno_id": turno.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def vista_cocina(request):
    """Pantalla de cocina para mostrar pedidos"""
    pedidos = PedidoCocina.objects.select_related('venta__mesa', 'articulo').exclude(
        estado='ENTREGADO'
    ).order_by('-fecha_creacion')
    return render(request, 'cocina.html', {'pedidos': pedidos})


@login_required
def api_pedido_cocina(request):
    if request.method == "POST":
        try:
            datos = json.loads(request.body)
            venta_id = datos.get('venta_id')
            items = datos.get('items', [])
            pedido_ids = []
            for item in items:
                articulo = Articulo.objects.get(id=item['id'])
                for _ in range(int(item['cantidad'])):
                    pc = PedidoCocina.objects.create(
                        venta_id=venta_id, articulo=articulo,
                        unidades=1, notas=item.get('notas', ''),
                    )
                    pedido_ids.append(pc.id)
            return JsonResponse({"status": "success", "pedidos": pedido_ids})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_validar_cupon(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        codigo = datos.get('codigo', '').strip().upper()
        subtotal = Decimal(str(datos.get('subtotal', 0)))
        cupon = CuponDescuento.objects.filter(codigo=codigo).first()
        if not cupon:
            return JsonResponse({"error": "Cupon no valido"}, status=400)
        if not cupon.esta_disponible:
            return JsonResponse({"error": "Cupon expirado o sin usos"}, status=400)
        descuento = cupon.calcular_descuento(subtotal)
        return JsonResponse({
            "status": "success", "descuento": float(descuento),
            "tipo": cupon.get_tipo_display(), "valor": float(cupon.valor),
            "descripcion": cupon.descripcion,
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_buscar_productos(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse([], safe=False)
    articulos = Articulo.objects.filter(
        Q(nombre__icontains=q) | Q(categoria__nombre__icontains=q),
        activo=True
    )[:20]
    results = []
    for a in articulos:
        results.append({
            'id': a.id, 'nombre': a.nombre,
            'precio': float(a.precio_con_iva),
            'categoria': a.categoria.nombre,
            'icono': a.categoria.icono,
        })
    return JsonResponse(results, safe=False)


@login_required
def api_top_vendidos(request):
    top = (
        LineaVenta.objects
        .values('articulo__id', 'articulo__nombre')
        .annotate(total=Sum('unidades'))
        .order_by('-total')[:10]
    )
    results = []
    for t in top:
        art = Articulo.objects.filter(id=t['articulo__id']).first()
        if art:
            results.append({
                'id': art.id, 'nombre': art.nombre,
                'precio': float(art.precio_con_iva),
                'icono': art.categoria.icono,
                'vendidos': int(t['total']),
            })
    return JsonResponse(results, safe=False)


def _generar_ticket_pdf(venta):
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    import io

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=(80*mm, 200*mm))
    w = 80*mm
    h = 200*mm
    y = h - 10*mm

    def text_center(txt, size=10, bold=False):
        nonlocal y
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
        c.drawCentredString(w/2, y, txt)
        y -= size + 2

    def text_line(txt, size=9):
        nonlocal y
        c.setFont('Helvetica', size)
        c.drawString(5*mm, y, txt)
        y -= size + 2

    text_center('TPV CAFETERIA', 14, True)
    text_center('================', 8)
    y -= 3*mm
    text_center(f'Ticket #{venta.id}', 10, True)
    text_center(venta.fecha_registro.strftime('%d/%m/%Y %H:%M'), 8)
    text_center(f'Cajero: {venta.empleado_caja.username}', 8)
    if venta.mesa:
        text_center(f'Mesa: {venta.mesa.numero}', 8)
    y -= 3*mm
    text_line('-' * 35, 8)

    for lv in venta.lineas.select_related('articulo'):
        subtotal = lv.precio_aplicado_con_iva * lv.unidades
        text_line(f'{lv.unidades}x {lv.articulo.nombre[:25]}', 9)
        text_line(f'  {subtotal:.2f} EUR', 9)

    text_line('-' * 35, 8)
    if venta.descuento_aplicado > 0:
        text_line(f'Descuento: -{venta.descuento_aplicado:.2f} EUR', 9)
        y -= 1*mm
    text_center(f'TOTAL: {venta.total_facturado:.2f} EUR', 11, True)
    text_center(f'({venta.forma_pago})', 8)
    y -= 3*mm
    text_line('-' * 35, 8)
    text_center('Gracias por su compra!', 9)
    y -= 5*mm
    text_line('-' * 35, 8)
    text_center('Comprobante verificado', 7)
    y -= 2*mm
    short_hash = venta.hash_seguridad[-8:].upper()
    text_center(f'Cod: {short_hash}', 7)

    c.save()
    buffer.seek(0)
    return buffer


@login_required
def pdf_ticket(request, venta_id):
    venta = OperacionVenta.objects.select_related('empleado_caja', 'mesa').get(id=venta_id)
    if not request.user.is_superuser and venta.empleado_caja != request.user:
        return HttpResponse("No autorizado", status=403)
    buffer = _generar_ticket_pdf(venta)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{venta.id}.pdf"'
    return response


@user_passes_test(lambda u: u.is_superuser)
def pdf_reporte_diario(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    import io

    from datetime import date as dt
    hoy = dt.today()
    ventas = OperacionVenta.objects.filter(fecha_registro__date=hoy).select_related('empleado_caja')
    total = ventas.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    ef = ventas.filter(forma_pago='EFECTIVO').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tj = ventas.filter(forma_pago='TARJETA').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    y = h - 40

    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(w/2, y, 'REPORTE DIARIO - TPV Cafeteria')
    y -= 25
    c.setFont('Helvetica', 11)
    c.drawCentredString(w/2, y, f'Fecha: {hoy.strftime("%d/%m/%Y")}')
    y -= 30

    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, 'RESUMEN')
    y -= 20
    c.setFont('Helvetica', 11)
    for label, val in [('Tickets totales', str(ventas.count())), ('Total facturado', f'{total:.2f} EUR'),
                       ('Efectivo', f'{ef:.2f} EUR'), ('Tarjeta', f'{tj:.2f} EUR')]:
        c.drawString(50, y, f'{label}:')
        c.drawRightString(w - 40, y, val)
        y -= 18

    y -= 15
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, 'DETALLE DE VENTAS')
    y -= 20
    c.setFont('Helvetica-Bold', 9)
    c.drawString(40, y, '#')
    c.drawString(60, y, 'Hora')
    c.drawString(100, y, 'Cajero')
    c.drawString(180, y, 'Pago')
    c.drawRightString(w - 40, y, 'Total')
    y -= 15
    c.line(40, y, w - 40, y)
    y -= 15

    c.setFont('Helvetica', 9)
    for v in ventas.order_by('fecha_registro'):
        if y < 50:
            c.showPage()
            y = h - 40
        c.drawString(40, y, str(v.id))
        c.drawString(60, y, v.fecha_registro.strftime('%H:%M'))
        c.drawString(100, y, v.empleado_caja.username[:12])
        c.drawString(180, y, v.forma_pago)
        c.drawRightString(w - 40, y, f'{v.total_facturado:.2f}')
        y -= 15

    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="reporte_{hoy}.pdf"'
    return response


@login_required
def vista_empleados(request):
    if not request.user.is_superuser:
        perfil = PerfilEmpleado.objects.filter(user=request.user).first()
        return render(request, 'empleados.html', {'mi_perfil': perfil, 'es_propio': True})
    empleados = PerfilEmpleado.objects.select_related('user').all()
    users_sin_perfil = User.objects.filter(perfil__isnull=True).exclude(username='admin')
    return render(request, 'empleados.html', {
        'empleados': empleados, 'users_sin_perfil': users_sin_perfil, 'es_propio': False,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def api_crear_empleado(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        user = User.objects.create_user(
            username=datos['username'], password=datos['password'],
            first_name=datos.get('nombre', ''), last_name=datos.get('apellidos', ''),
        )
        perfil = PerfilEmpleado.objects.create(
            user=user, rol=datos.get('rol', 'CAJERO'),
            telefono=datos.get('telefono', ''), pin_acceso=datos.get('pin', ''),
        )
        AuditLog.objects.create(
            usuario=request.user, accion='CREAR', modelo='Empleado',
            objeto_id=user.id,
            descripcion=f"Empleado creado: {user.username} ({perfil.get_rol_display()})",
            ip_address=get_ip(request),
        )
        return JsonResponse({"status": "success", "user_id": user.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_actualizar_pedido_cocina(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        pedido_id = datos.get('pedido_id')
        nuevo_estado = datos.get('estado')
        pedido = PedidoCocina.objects.get(id=pedido_id)
        pedido.estado = nuevo_estado
        if nuevo_estado == 'LISTO':
            pedido.fecha_listo = timezone.now()
        pedido.save()
        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def vista_menu_dia(request):
    hoy = date.today()
    ahora = datetime.now().time()
    if time(8, 0) <= ahora < time(12, 0):
        franja_actual = 'MANANA'
    elif time(12, 0) <= ahora < time(16, 0):
        franja_actual = 'MEDIODIA'
    elif time(16, 0) <= ahora < time(20, 0):
        franja_actual = 'TARDE'
    else:
        franja_actual = 'TODO'

    promos = MenuDelDia.objects.filter(
        activo=True, fecha=hoy
    ).filter(Q(franja=franja_actual) | Q(franja='TODO')).select_related('articulo')

    todos = MenuDelDia.objects.filter(activo=True).select_related('articulo')
    return render(request, 'menu_dia.html', {
        'promos': promos,
        'todos': todos,
        'franja_actual': franja_actual,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def vista_reservas(request):
    from datetime import timedelta
    hoy = date.today()
    semana = [hoy + timedelta(days=i) for i in range(7)]
    reservas_hoy = Reserva.objects.filter(
        fecha_reserva=hoy, estado__in=['PENDIENTE', 'CONFIRMADA']
    ).select_related('mesa')
    todas_reservas = Reserva.objects.filter(
        fecha_reserva__in=semana
    ).select_related('mesa').order_by('fecha_reserva', 'hora_reserva')
    mesas = Mesa.objects.filter(activa=True).order_by('numero')
    return render(request, 'reservas.html', {
        'reservas_hoy': reservas_hoy,
        'todas_reservas': todas_reservas,
        'semana': semana,
        'hoy': hoy,
        'mesas': mesas,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def api_crear_reserva(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        mesa = Mesa.objects.get(id=datos['mesa_id'])
        reserva = Reserva.objects.create(
            mesa=mesa, cliente_nombre=datos['nombre'],
            cliente_telefono=datos.get('telefono', ''),
            fecha_reserva=datos['fecha'], hora_reserva=datos['hora'],
            num_personas=int(datos.get('personas', 2)),
            notas=datos.get('notas', ''),
            creado_por=request.user,
        )
        AuditLog.objects.create(
            usuario=request.user, accion='CREAR', modelo='Reserva',
            objeto_id=reserva.id,
            descripcion=f"Reserva: {reserva.cliente_nombre} Mesa {mesa.numero} {reserva.fecha_reserva} {reserva.hora_reserva}",
            ip_address=get_ip(request),
        )
        return JsonResponse({"status": "success", "reserva_id": reserva.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def api_estado_reserva(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        reserva = Reserva.objects.get(id=datos['reserva_id'])
        reserva.estado = datos['estado']
        reserva.save()
        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_recomendaciones(request):
    articulo_id = request.GET.get('articulo_id')
    if not articulo_id:
        hoy = date.today()
        ventas_hoy = OperacionVenta.objects.filter(fecha_registro__date=hoy)
        vendidos = (
            LineaVenta.objects.filter(venta__in=ventas_hoy)
            .values('articulo__id', 'articulo__nombre', 'articulo__precio_sin_iva')
            .annotate(total=Sum('unidades'))
            .order_by('-total')[:5]
        )
        results = []
        for v in vendidos:
            art = Articulo.objects.filter(id=v['articulo__id'], activo=True).first()
            if art:
                cat = art.categoria
                results.append({
                    'id': art.id, 'nombre': art.nombre,
                    'precio': float(art.precio_con_iva),
                    'icono': cat.icono if cat else '☕',
                    'motivo': f'Mas vendido hoy ({v["total"]} uds)',
                })
        return JsonResponse(results, safe=False)

    ventas_con_producto = LineaVenta.objects.filter(
        articulo_id=articulo_id
    ).values_list('venta_id', flat=True)[:200]
    companions = (
        LineaVenta.objects.filter(venta_id__in=ventas_con_producto)
        .exclude(articulo_id=articulo_id)
        .values('articulo__id', 'articulo__nombre', 'articulo__precio_sin_iva')
        .annotate(freq=Count('id'))
        .order_by('-freq')[:5]
    )
    results = []
    for c in companions:
        art = Articulo.objects.filter(id=c['articulo__id'], activo=True).first()
        if art:
            cat = art.categoria
            results.append({
                'id': art.id, 'nombre': art.nombre,
                'precio': float(art.precio_con_iva),
                'icono': cat.icono if cat else '☕',
                'motivo': f'Comprado junto ({c["freq"]} veces)',
            })
    return JsonResponse(results, safe=False)


@login_required
def api_metas_hoy(request):
    hoy = date.today()
    meta, _ = MetaDiaria.objects.get_or_create(fecha=hoy, defaults={
        'objetivo_tickets': 50, 'objetivo_ingresos': Decimal('500.00')
    })

    ranking = (
        OperacionVenta.objects.filter(fecha_registro__date=hoy)
        .values('empleado_caja__username')
        .annotate(
            total_ventas=Count('id'),
            total_ingresos=Sum('total_facturado')
        )
        .order_by('-total_ingresos')[:10]
    )
    ranking_list = []
    for i, r in enumerate(ranking):
        ranking_list.append({
            'posicion': i + 1,
            'usuario': r['empleado_caja__username'],
            'ventas': r['total_ventas'],
            'ingresos': float(r['total_ingresos'] or 0),
        })

    return JsonResponse({
        'meta': {
            'objetivo_tickets': meta.objetivo_tickets,
            'objetivo_ingresos': float(meta.objetivo_ingresos),
            'tickets_hoy': meta.tickets_hoy,
            'ingresos_hoy': float(meta.ingresos_hoy),
            'porcentaje_tickets': meta.porcentaje_tickets,
            'porcentaje_ingresos': meta.porcentaje_ingresos,
            'porcentaje_total': meta.porcentaje_total,
            'badges': meta.badges,
        },
        'ranking': ranking_list,
    })


@login_required
def vista_pantalla_cliente(request):
    return render(request, 'pantalla_cliente.html')


@login_required
def api_estado_pedido_cliente(request):
    mesa_id = request.GET.get('mesa_id')
    if not mesa_id:
        return JsonResponse([], safe=False)
    from datetime import timedelta
    hace_2h = timezone.now() - timedelta(hours=2)
    pedidos = PedidoCocina.objects.filter(
        venta__mesa_id=mesa_id,
        fecha_creacion__gte=hace_2h
    ).select_related('articulo').order_by('-fecha_creacion')[:20]
    items = []
    for p in pedidos:
        items.append({
            'articulo': p.articulo.nombre,
            'icono': p.articulo.categoria.icono if p.articulo.categoria else '☕',
            'unidades': p.unidades,
            'estado': p.estado,
            'notas': p.notas,
            'hora': p.fecha_creacion.strftime('%H:%M'),
        })
    return JsonResponse(items, safe=False)


def pdf_ticket_con_qr(request, venta_id):
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    import io
    import qrcode
    from io import BytesIO

    venta = OperacionVenta.objects.select_related('empleado_caja', 'mesa').get(id=venta_id)
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=(80*mm, 200*mm))
    w = 80*mm
    h = 200*mm
    y = h - 10*mm

    def text_center(txt, size=10, bold=False):
        nonlocal y
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
        c.drawCentredString(w/2, y, txt)
        y -= size + 2

    def text_line(txt, size=9):
        nonlocal y
        c.setFont('Helvetica', size)
        c.drawString(5*mm, y, txt)
        y -= size + 2

    text_center('TPV CAFETERIA', 14, True)
    text_center('================', 8)
    y -= 3*mm
    text_center(f'Ticket #{venta.id}', 10, True)
    text_center(venta.fecha_registro.strftime('%d/%m/%Y %H:%M'), 8)
    text_center(f'Cajero: {venta.empleado_caja.username}', 8)
    if venta.mesa:
        text_center(f'Mesa: {venta.mesa.numero}', 8)
    y -= 3*mm
    text_line('-' * 35, 8)

    for lv in venta.lineas.select_related('articulo'):
        subtotal = lv.precio_aplicado_con_iva * lv.unidades
        text_line(f'{lv.unidades}x {lv.articulo.nombre[:25]}', 9)
        text_line(f'  {subtotal:.2f} EUR', 9)

    text_line('-' * 35, 8)
    if venta.descuento_aplicado > 0:
        text_line(f'Descuento: -{venta.descuento_aplicado:.2f} EUR', 9)
        y -= 1*mm
    text_center(f'TOTAL: {venta.total_facturado:.2f} EUR', 11, True)
    text_center(f'({venta.forma_pago})', 8)
    y -= 3*mm
    text_line('-' * 35, 8)
    text_center('Gracias por su compra!', 9)
    y -= 5*mm
    text_line('-' * 35, 8)
    text_center('Comprobante verificado', 7)
    y -= 2*mm
    short_hash = venta.hash_seguridad[-8:].upper()
    text_center(f'Cod: {short_hash}', 7)

    y -= 8*mm
    qr_data = f'TPV#{venta.id}|{venta.total_facturado}|{venta.forma_pago}|{venta.fecha_registro.strftime("%d%m%Y%H%M")}|{venta.hash_seguridad[:16]}'
    qr = qrcode.make(qr_data, box_size=3, border=1)
    qr_buf = BytesIO()
    qr.save(qr_buf, format='PNG')
    qr_buf.seek(0)
    from reportlab.lib.utils import ImageReader
    qr_img = ImageReader(qr_buf)
    qr_size = 25*mm
    c.drawImage(qr_img, (w - qr_size) / 2, y - qr_size, qr_size, qr_size)
    y -= qr_size + 3*mm
    text_center('Escanea para ver tu ticket digital', 7)

    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{venta.id}.pdf"'
    return response


@login_required
def api_buscar_codigo_barras(request):
    codigo = request.GET.get('codigo', '').strip()
    if not codigo:
        return JsonResponse({"error": "Codigo vacio"}, status=400)
    articulo = Articulo.objects.filter(codigo_barras=codigo, activo=True).first()
    if not articulo:
        return JsonResponse({"error": "Producto no encontrado"}, status=404)
    return JsonResponse({
        'id': articulo.id, 'nombre': articulo.nombre,
        'precio': float(articulo.precio_con_iva),
        'icono': articulo.categoria.icono if articulo.categoria else '☕',
        'categoria': articulo.categoria.nombre if articulo.categoria else '',
    })


@login_required
def api_split_bill(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        venta_id = datos.get('venta_id')
        participantes = datos.get('participantes', [])
        venta = OperacionVenta.objects.get(id=venta_id)
        total = venta.total_facturado
        if not participantes:
            return JsonResponse({"error": "Sin participantes"}, status=400)
        n = len(participantes)
        por_persona = (total / Decimal(str(n))).quantize(Decimal('0.01'))
        restante = total - (por_persona * Decimal(str(n - 1)))
        results = []
        for i, p in enumerate(participantes):
            monto = por_persona if i < n - 1 else restante
            results.append({
                'nombre': p.get('nombre', f'Persona {i+1}'),
                'metodo': p.get('metodo', 'EFECTIVO'),
                'monto': float(monto),
            })
        return JsonResponse({"status": "success", "particiones": results, "total": float(total)})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_clientes(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse([], safe=False)
    clientes = Cliente.objects.filter(
        Q(nombre__icontains=q) | Q(email__icontains=q) | Q(telefono__icontains=q),
        activo=True
    )[:10]
    results = []
    for cl in clientes:
        results.append({
            'id': cl.id, 'nombre': cl.nombre, 'email': cl.email,
            'nivel': cl.get_nivel_display(), 'puntos': cl.puntos,
            'descuento': cl.descuento_nivel,
        })
    return JsonResponse(results, safe=False)


@login_required
def api_crear_cliente(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        cl = Cliente.objects.create(
            nombre=datos['nombre'], email=datos.get('email', ''),
            telefono=datos.get('telefono', ''),
        )
        return JsonResponse({"status": "success", "cliente_id": cl.id, "nombre": cl.nombre})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_transferir_mesa(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        origen_id = datos.get('mesa_origen_id')
        destino_id = datos.get('mesa_destino_id')
        destino = Mesa.objects.get(id=destino_id)
        ventas = OperacionVenta.objects.filter(mesa_id=origen_id, total_facturado=0)
        ventas.update(mesa=destino)
        return JsonResponse({"status": "success", "mensaje": f"Transferido a Mesa {destino.numero}"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_timer_cocina(request):
    if request.method == "GET":
        pedidos = PedidoCocina.objects.filter(
            estado__in=['PENDIENTE', 'PREPARANDO']
        ).select_related('articulo', 'venta', 'venta__mesa').order_by('fecha_creacion')
        items = []
        for p in pedidos:
            elapsed = (timezone.now() - p.fecha_creacion).total_seconds()
            items.append({
                'id': p.id,
                'articulo': p.articulo.nombre,
                'icono': p.articulo.categoria.icono if p.articulo.categoria else '☕',
                'mesa': p.venta.mesa.numero if p.venta.mesa else 'Llevar',
                'unidades': p.unidades,
                'estado': p.estado,
                'elapsed_seconds': int(elapsed),
                'fecha_creacion': p.fecha_creacion.isoformat(),
            })
        return JsonResponse(items, safe=False)
    elif request.method == "POST":
        try:
            datos = json.loads(request.body)
            pedido = PedidoCocina.objects.get(id=datos['pedido_id'])
            pedido.estado = datos['estado']
            if datos['estado'] == 'LISTO':
                pedido.fecha_listo = timezone.now()
            pedido.save()
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_clima_sugerencias(request):
    import urllib.request
    try:
        req = urllib.request.Request(
            'https://wttr.in/?format=j1&lang=es',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        current = data.get('current_condition', [{}])[0]
        temp = int(current.get('temp_C', 20))
        desc_es = current.get('lang_es', [{}])[0].get('value', current.get('weatherDesc', [{}])[0].get('value', ''))
        if temp < 10:
            suggest_cat = 'Bebidas Calientes'
            emoji = '🥶'
        elif temp < 22:
            suggest_cat = None
            emoji = '🌤'
        else:
            suggest_cat = 'Bebidas Frias'
            emoji = '🔥'
        sugerencias = []
        if suggest_cat:
            arts = Articulo.objects.filter(categoria__nombre__icontains=suggest_cat, activo=True)[:4]
            for a in arts:
                sugerencias.append({'nombre': a.nombre, 'precio': float(a.precio_con_iva), 'icono': a.categoria.icono})
        return JsonResponse({
            'temp': temp, 'desc': desc_es, 'emoji': emoji,
            'sugerencias': sugerencias,
        })
    except Exception:
        return JsonResponse({'temp': 20, 'desc': 'No disponible', 'emoji': '☀️', 'sugerencias': []})


@login_required
def vista_heatmap(request):
    from django.db.models.functions import TruncHour, TruncDay
    ventas = OperacionVenta.objects.all()
    por_hora = (
        ventas.annotate(hora=TruncHour('fecha_registro'))
        .values('hora')
        .annotate(total=Count('id'), ingresos=Sum('total_facturado'))
        .order_by('hora')
    )
    por_dia_raw = (
        ventas.annotate(dia=TruncDay('fecha_registro'))
        .values('dia')
        .annotate(total=Count('id'), ingresos=Sum('total_facturado'))
        .order_by('dia')
    )
    from collections import defaultdict
    dias_nombres = {0: 'Lun', 1: 'Mar', 2: 'Mie', 3: 'Jue', 4: 'Vie', 5: 'Sab', 6: 'Dom'}
    por_dia = defaultdict(lambda: {'total': 0, 'ingresos': 0})
    for item in por_dia_raw:
        dw = item['dia'].weekday()
        por_dia[dw]['total'] += item['total'] or 0
        por_dia[dw]['ingresos'] = float(por_dia[dw]['ingresos']) + float(item['ingresos'] or 0)
    por_dia_list = [{'dia_semana': k, 'dia_nombre': dias_nombres[k], 'total': v['total'], 'ingresos': v['ingresos']} for k, v in sorted(por_dia.items())]
    return render(request, 'heatmap.html', {
        'por_hora': list(por_hora),
        'por_dia': por_dia_list,
    })


@login_required
def api_forecast(request):
    from datetime import timedelta
    hoy = date.today()
    forecast_dias = []
    for i in range(1, 8):
        dia_futuro = hoy + timedelta(days=i)
        dia_semana = dia_futuro.weekday()
        historico = OperacionVenta.objects.filter(
            fecha_registro__month=dia_futuro.month,
            fecha_registro__weekday=dia_semana
        )
        tickets_avg = historico.count() / max(1, historico.dates('fecha_registro', 'day').count())
        ingresos_avg_data = historico.aggregate(avg=Sum('total_facturado'))
        ingresos_avg = float(ingresos_avg_data['avg'] or 0) / max(1, historico.dates('fecha_registro', 'day').count())
        forecast_dias.append({
            'fecha': dia_futuro.strftime('%d/%m'),
            'dia_nombre': ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom'][dia_semana],
            'tickets_est': round(tickets_avg * (1.1 if dia_semana >= 5 else 1.0)),
            'ingresos_est': round(ingresos_avg * (1.15 if dia_semana >= 5 else 1.0), 2),
        })

    historico_hoy = OperacionVenta.objects.filter(fecha_registro__date=hoy)
    hoy_tickets = historico_hoy.count()
    hoy_ingresos = float(historico_hoy.aggregate(t=Sum('total_facturado'))['t'] or 0)
    semana_tickets = OperacionVenta.objects.filter(fecha_registro__gte=hoy - timedelta(days=7)).count()
    semana_ingresos = float(OperacionVenta.objects.filter(
        fecha_registro__gte=hoy - timedelta(days=7)
    ).aggregate(t=Sum('total_facturado'))['t'] or 0)

    return JsonResponse({
        'hoy': {'tickets': hoy_tickets, 'ingresos': round(hoy_ingresos, 2)},
        'semana': {'tickets': semana_tickets, 'ingresos': round(semana_ingresos, 2)},
        'forecast': forecast_dias,
    })


@login_required
def vista_plano(request):
    plano, _ = PlanoRestaurante.objects.get_or_create(activo=True, defaults={'nombre': 'Plano Principal'})
    posiciones = MesaPosicion.objects.filter(plano=plano).select_related('mesa')
    posiciones_mesa = [p.mesa for p in posiciones]
    mesas = Mesa.objects.filter(activa=True).order_by('numero')
    return render(request, 'plano.html', {
        'plano': plano, 'posiciones': posiciones, 'posiciones_mesa': posiciones_mesa, 'mesas': mesas,
    })


@login_required
def api_guardar_posiciones(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        posiciones = datos.get('posiciones', [])
        plano, _ = PlanoRestaurante.objects.get_or_create(activo=True, defaults={'nombre': 'Plano Principal'})
        for pos in posiciones:
            mesa = Mesa.objects.get(id=pos['mesa_id'])
            MesaPosicion.objects.update_or_create(
                plano=plano, mesa=mesa,
                defaults={'x': pos['x'], 'y': pos['y'], 'color': pos.get('color', '#00d4ff')}
            )
        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def vista_pedido_movil(request, mesa_numero):
    try:
        mesa = Mesa.objects.get(numero=mesa_numero, activa=True)
    except Mesa.DoesNotExist:
        return render(request, 'pedido_movil.html', {'error': 'Mesa no encontrada'})
    articulos = Articulo.objects.filter(activo=True).select_related('categoria')
    categorias = CategoriaProducto.objects.all()
    return render(request, 'pedido_movil.html', {
        'mesa': mesa, 'articulos': articulos, 'categorias': categorias,
    })


def api_pedido_movil_crear(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        mesa = Mesa.objects.get(id=datos['mesa_id'])
        pedido = PedidoMovil.objects.create(
            mesa=mesa, cliente_nombre=datos.get('nombre', ''),
            notas=datos.get('notas_general', ''),
        )
        for item in datos.get('items', []):
            art = Articulo.objects.get(id=item['id'])
            LineaPedidoMovil.objects.create(
                pedido=pedido, articulo=art,
                unidades=item.get('cantidad', 1),
                notas=item.get('notas', ''),
            )
        AuditLog.objects.create(
            usuario=None, accion='CREAR', modelo='PedidoMovil',
            objeto_id=pedido.id,
            descripcion=f"Pedido movil Mesa {mesa.numero}: {len(datos.get('items', []))} items",
        )
        return JsonResponse({"status": "success", "pedido_id": pedido.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def api_pedidos_movil_pendientes(request):
    pedidos = PedidoMovil.objects.filter(
        estado__in=['RECIBIDO', 'CONFIRMADO', 'PREPARANDO']
    ).select_related('mesa').order_by('fecha')
    items = []
    for p in pedidos:
        lineas = []
        for l in p.lineas.select_related('articulo'):
            lineas.append({
                'articulo': l.articulo.nombre,
                'icono': l.articulo.categoria.icono if l.articulo.categoria else '☕',
                'unidades': l.unidades,
                'notas': l.notas,
            })
        items.append({
            'id': p.id,
            'mesa': p.mesa.numero,
            'estado': p.estado,
            'cliente': p.cliente_nombre,
            'notas': p.notas,
            'fecha': p.fecha.strftime('%H:%M'),
            'lineas': lineas,
        })
    return JsonResponse(items, safe=False)


@login_required
def api_social_proof(request):
    from datetime import timedelta
    hace_30 = timezone.now() - timedelta(minutes=30)
    ventas_recientes = OperacionVenta.objects.filter(
        fecha_registro__gte=hace_30
    ).select_related('empleado_caja', 'mesa').order_by('-fecha_registro')[:5]
    items = []
    for v in ventas_recientes:
        nombres = []
        for lv in v.lineas.select_related('articulo')[:2]:
            nombres.append(lv.articulo.nombre)
        mesa_text = f"Mesa {v.mesa.numero}" if v.mesa else "Para llevar"
        items.append({
            'cliente': v.cliente_nombre if hasattr(v, 'cliente_nombre') and v.cliente_nombre else mesa_text,
            'productos': nombres,
            'total': float(v.total_facturado),
            'hora': v.fecha_registro.strftime('%H:%M'),
        })
    return JsonResponse(items, safe=False)


@login_required
def api_satisfaccion(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        venta = OperacionVenta.objects.get(id=datos['venta_id'])
        venta.satisfaccion = datos['rating']
        venta.save()
        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_stats_satisfaccion(request):
    total = OperacionVenta.objects.exclude(satisfaccion='').count()
    if total == 0:
        return JsonResponse({'total': 0, 'excelente': 0, 'buena': 0, 'normal': 0, 'mala': 0, 'promedio': 0})
    stats = {}
    for val, emoji in OperacionVenta._meta.get_field('satisfaccion').choices:
        count = OperacionVenta.objects.filter(satisfaccion=val).count()
        stats[val.lower()] = count
    promedio = (stats.get('excelente', 0) * 4 + stats.get('buena', 0) * 3 +
                stats.get('normal', 0) * 2 + stats.get('mala', 0) * 1) / total
    return JsonResponse({
        'total': total,
        'excelente': stats.get('excelente', 0),
        'buena': stats.get('buena', 0),
        'normal': stats.get('normal', 0),
        'mala': stats.get('mala', 0),
        'promedio': round(promedio, 2),
    })


@login_required
def api_cola_siguiente(request):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo no permitido"}, status=405)
    try:
        datos = json.loads(request.body)
        accion = datos.get('accion', 'siguiente')
        if accion == 'siguiente':
            num = ColaNumero.objects.filter(estado='ESPERANDO', fecha=date.today()).first()
            if num:
                num.estado = 'LLAMANDO'
                num.atendido = timezone.now()
                num.save()
                return JsonResponse({"status": "success", "numero": num.numero, "nombre": num.nombre_cliente})
            return JsonResponse({"status": "empty", "mensaje": "No hay nadie en la cola"})
        elif accion == 'atender':
            num = ColaNumero.objects.filter(estado='LLAMANDO', fecha=date.today()).first()
            if num:
                num.estado = 'ATENDIDO'
                num.save()
                return JsonResponse({"status": "success"})
        elif accion == 'nuevo':
            ultimo = ColaNumero.objects.filter(fecha=date.today()).order_by('-numero').first()
            nuevo_num = (ultimo.numero + 1) if ultimo else 1
            cola = ColaNumero.objects.create(
                numero=nuevo_num,
                nombre_cliente=datos.get('nombre', ''),
            )
            return JsonResponse({"status": "success", "numero": cola.numero})
        elif accion == 'cancelar':
            ColaNumero.objects.filter(estado__in=['ESPERANDO', 'LLAMANDO'], fecha=date.today()).update(estado='CANCELADO')
            return JsonResponse({"status": "success"})
        return JsonResponse({"error": "Accion no valida"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def api_cola_estado(request):
    esperando = ColaNumero.objects.filter(estado='ESPERANDO', fecha=date.today()).count()
    llamando = ColaNumero.objects.filter(estado='LLAMANDO', fecha=date.today()).first()
    ultimo = ColaNumero.objects.filter(fecha=date.today()).order_by('-numero').first()
    return JsonResponse({
        'esperando': esperando,
        'llamando': llamando.numero if llamando else None,
        'llamando_nombre': llamando.nombre_cliente if llamando else None,
        'ultimo_numero': ultimo.numero if ultimo else 0,
    })


@login_required
def vista_cola(request):
    return render(request, 'cola.html')


# ==================== 8 NUEVAS FEATURES ====================

# 1. VOZ AUTOMATICA - TTS anuncia pedidos a cocina
@login_required
@user_passes_test(es_empleado)
def api_anuncio_voz(request):
    """Ultimo pedido listo para que la voz lo anuncie"""
    pedido = PedidoCocina.objects.filter(estado='LISTO').order_by('-fecha_listo').first()
    if not pedido:
        return JsonResponse({'mensaje': ''})
    linea = pedido.lineas.first()
    articulo = linea.articulo.nombre if linea else 'Pedido'
    mesa_text = f"Mesa {pedido.mesa.numero}" if pedido.mesa else "Para llevar"
    texto = f"{mesa_text}, tu {articulo} está listo. Por favor pásalo a recoger."
    pedido.announcement = True
    pedido.save()
    return JsonResponse({
        'mensaje': texto,
        'mesa': pedido.mesa.numero if pedido.mesa else None,
        'articulo': articulo,
    })


# 2. ANIMACIONES 3D - productos al añadir al ticket
@login_required
@user_passes_test(es_empleado)
def api_producto_animacion(request, producto_id):
    """Datos del producto para animacion 3D al agregar"""
    try:
        art = Articulo.objects.get(id=producto_id)
        return JsonResponse({
            'nombre': art.nombre,
            'precio': float(art.precio_con_iva),
            'imagen': '/static/img/cafe_default.svg',
        })
    except Articulo.DoesNotExist:
        return JsonResponse({'error': 'No encontrado'}, status=404)


# 3. COMPETICION CAJEROS - game化
@login_required
@user_passes_test(es_gerente)
def api_competicion(request):
    """Ranking de cajeros por ventas del mes actual"""
    hoy = timezone.now().replace(day=1).date()
    ranking = (
        OperacionVenta.objects
        .filter(fecha_registro__date__gte=hoy)
        .values('empleado_caja__username', 'empleado_caja__first_name')
        .annotate(
            ventas=Count('id'),
            ingresos=Sum('total_facturado'),
        )
        .order_by('-ingresos')[:10]
    )
    badges = []
    for i, r in enumerate(ranking):
        badge = {'oro': 0, 'plata': 0, 'bronce': 0}
        if r['ventas'] >= 100:
            badge['oro'] = 1
        elif r['ventas'] >= 50:
            badge['plata'] = 1
        elif r['ventas'] >= 20:
            badge['bronce'] = 1
        badges.append({
            'posicion': i + 1,
            'username': r['empleado_caja__username'],
            'nombre': r['empleado_caja__first_name'] or r['empleado_caja__username'],
            'ventas': r['ventas'],
            'ingresos': float(r['ingresos'] or 0),
            'badge': badge,
        })
    return JsonResponse({'ranking': badges})


# 4. SUGERENCIA INTELIGENTE POR HORA
@login_required
@user_passes_test(es_empleado)
def api_sugerencia_hora(request):
    """Sugiere productos segun hora del dia"""
    ahora = timezone.localtime(timezone.now())
    hora = ahora.hour
    if 6 <= hora < 10:
        sugerencia = "Buenos dias! La hora perfecta para un cafe con leche y medialunas"
        productos_ids = Articulo.objects.filter(nombre__icontains='leche').values_list('id', flat=True)[:3]
    elif 10 <= hora < 13:
        sugerencia = "Media manana? Un te con algo dulce viene genial"
        productos_ids = Articulo.objects.filter(categoria__nombre__icontains='te').values_list('id', flat=True)[:3]
    elif 13 <= hora < 16:
        sugerencia = "Hora de almuerzo! Revisa nuestro menu del dia"
        productos_ids = MenuDelDia.objects.filter(activo=True).values_list('articulo_id', flat=True)[:3]
    elif 16 <= hora < 19:
        sugerencia = "Merienda! Un espresso o capuchino con pastel"
        productos_ids = Articulo.objects.filter(nombre__icontains='capuchino').values_list('id', flat=True)[:3]
    else:
        sugerencia = "Noche de cafe? Un espresso doble para llevar"
        productos_ids = Articulo.objects.filter(nombre__icontains='espresso').values_list('id', flat=True)[:3]
    productos = Articulo.objects.filter(id__in=productos_ids).values('id', 'nombre', 'precio_sin_iva')
    result = [{'id': p['id'], 'nombre': p['nombre'], 'precio': float(p['precio_sin_iva'])} for p in productos]
    return JsonResponse({
        'sugerencia': sugerencia,
        'hora': ahora.strftime('%H:%M'),
        'productos': result,
    })


# 5. AUTO-PEDIDO PROVEEDORES
@login_required
@user_passes_test(es_gerente)
def api_stock_bajo(request):
    """Lista insumos con stock bajo"""
    criticos = InsumoMateriaPrima.objects.filter(cantidad_actual__lte=F('cantidad_minima'))
    lista = []
    for i in criticos:
        lista.append({
            'id': i.id,
            'nombre': i.nombre,
            'unidad': i.unidad_medida,
            'stock_actual': float(i.cantidad_actual),
            'stock_minimo': float(i.cantidad_minima),
            'faltante': float(i.cantidad_minima - i.cantidad_actual),
        })
    return JsonResponse({'insumos': lista})


@login_required
@user_passes_test(es_gerente)
def api_generar_pedido_proveedor(request):
    """Genera automaticamente pedidos de compra para insumos criticos"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    criticos = InsumoMateriaPrima.objects.filter(cantidad_actual__lte=F('cantidad_minima'))
    pedidos = []
    for i in criticos:
        cantidad = i.cantidad_minima * 2  # Pedir el doble del minimo
        pedido = PedidoProveedor.objects.create(
            insumo=i,
            cantidad_solicitada=cantidad,
            notas='Auto-generado por stock bajo',
            creado_por=request.user,
        )
        pedidos.append({
            'insumo': i.nombre,
            'cantidad': float(cantidad),
            'unidad': i.unidad_medida,
        })
    return JsonResponse({'pedidos_creados': pedidos, 'total': len(pedidos)})


# 6. CHAT CAJERO-COCINA
@login_required
@user_passes_test(es_empleado)
def api_chat_mensajes(request):
    """Obtener mensajes del chat"""
    desde = request.GET.get('desde', '')
    qs = MensajeChat.objects.all()
    if desde:
        qs = qs.filter(fecha__gt=desde)
    mensajes = []
    for m in qs.order_by('-fecha')[:50]:
        mensajes.append({
            'id': m.id,
            'emisor': m.emisor,
            'texto': m.texto,
            'fecha': m.fecha.strftime('%H:%M'),
            'leido': m.leido,
        })
    return JsonResponse({'mensajes': list(reversed(mensajes))})


@login_required
@user_passes_test(es_empleado)
def api_chat_enviar(request):
    """Enviar mensaje al chat"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    emisor = 'COCINA' if es_cocina(request.user) else 'CAJERO'
    MensajeChat.objects.create(emisor=emisor, texto=data.get('texto', ''))
    return JsonResponse({'ok': True})


# 7. PANTALLA ESCAPARATE 3D
@login_required
@user_passes_test(es_gerente)
def vista_escaparate(request):
    """Pantalla 3D para mostrar en el escaparate de la tienda"""
    articulos = Articulo.objects.filter(activo=True).order_by('nombre')
    productos_json = json.dumps([{
        'nombre': a.nombre,
        'precio': float(a.precio_con_iva),
    } for a in articulos])
    return render(request, 'escaparate.html', {'productos': productos_json})


# 8. EXPORT EXCEL/CSV COMPLETO
@login_required
@user_passes_test(es_gerente)
def api_exportar_excel(request):
    """Exporta ventas del dia a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # Hoja 1: Ventas del dia
    ws = wb.active
    ws.title = "Ventas del Dia"
    headers = ['ID', 'Fecha', 'Cajero', 'Mesa', 'Subtotal', 'IVA', 'Total', 'Metodo Pago']
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    hoy = timezone.now().date()
    ventas = OperacionVenta.objects.filter(fecha_registro__date=hoy)
    for row, v in enumerate(ventas, 2):
        ws.cell(row=row, column=1, value=v.id)
        ws.cell(row=row, column=2, value=v.fecha_registro.strftime('%d/%m %H:%M'))
        ws.cell(row=row, column=3, value=str(v.empleado_caja))
        ws.cell(row=row, column=4, value=f"Mesa {v.mesa.numero}" if v.mesa else 'N/A')
        ws.cell(row=row, column=5, value=float(v.subtotal_base))
        ws.cell(row=row, column=6, value=float(v.total_impuestos))
        ws.cell(row=row, column=7, value=float(v.total_facturado))
        ws.cell(row=row, column=8, value=v.forma_pago)

    for col in range(1, 9):
        ws.column_dimensions[chr(64+col)].width = 18

    # Hoja 2: Top productos
    ws2 = wb.create_sheet("Top Productos")
    top_headers = ['Producto', 'Unidades Vendidas', 'Ingresos']
    for col, h in enumerate(top_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")

    top = (
        LineaVenta.objects
        .filter(venta__fecha_registro__date=hoy)
        .values('articulo__nombre')
        .annotate(unidades=Sum('unidades'), ingresos=Sum('precio_aplicado_con_iva'))
        .order_by('-ingresos')[:20]
    )
    for row, t in enumerate(top, 2):
        ws2.cell(row=row, column=1, value=t['articulo__nombre'])
        ws2.cell(row=row, column=2, value=t['unidades'])
        ws2.cell(row=row, column=3, value=float(t['ingresos'] or 0))

    for col in range(1, 4):
        ws2.column_dimensions[chr(64+col)].width = 22

    # Hoja 3: Gastos
    ws3 = wb.create_sheet("Gastos")
    gastos_headers = ['Fecha', 'Concepto', 'Monto', 'Categoria']
    for col, h in enumerate(gastos_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")

    gastos = RegistroGasto.objects.filter(fecha_gasto__date=hoy)
    for row, g in enumerate(gastos, 2):
        ws3.cell(row=row, column=1, value=g.fecha_gasto.strftime('%H:%M'))
        ws3.cell(row=row, column=2, value=g.concepto)
        ws3.cell(row=row, column=3, value=float(g.importe_total))
        ws3.cell(row=row, column=4, value='General')

    for col in range(1, 5):
        ws3.column_dimensions[chr(64+col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tpv_reporte_{hoy.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ==================== 8 NUEVAS FEATURES (ROUND 4) ====================

# 1. FIRMA DIGITAL EN PANTALLA
@login_required
@user_passes_test(es_empleado)
def api_guardar_firma(request):
    """Guarda la firma digital del cliente para un ticket"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    venta_id = data.get('venta_id')
    firma_base64 = data.get('firma', '')
    nombre = data.get('nombre', '')
    try:
        venta = OperacionVenta.objects.get(id=venta_id)
        FirmaDigital.objects.update_or_create(
            venta=venta,
            defaults={'imagen_firma': firma_base64, 'cliente_nombre': nombre}
        )
        return JsonResponse({'ok': True})
    except OperacionVenta.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)


# 2. KIOSK AUTO-SERVICIO
@login_required
def vista_kiosk(request):
    """Pantalla de auto-servicio para clientes"""
    categorias = CategoriaProducto.objects.all()
    articulos = Articulo.objects.filter(activo=True)
    mesas = Mesa.objects.filter(activa=True).order_by('numero')
    return render(request, 'kiosk.html', {
        'categorias': categorias,
        'articulos': articulos,
        'mesas': mesas,
    })


@login_required
@user_passes_test(es_empleado)
def api_kiosk_pedido(request):
    """Crea pedido desde kiosk auto-servicio"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    mesa_id = data.get('mesa_id')
    items = data.get('items', [])
    if not mesa_id or not items:
        return JsonResponse({'error': 'Faltan datos'}, status=400)
    try:
        mesa = Mesa.objects.get(id=mesa_id)
    except Mesa.DoesNotExist:
        return JsonResponse({'error': 'Mesa no encontrada'}, status=404)
    try:
        cajero = User.objects.get(username='admin')
    except User.DoesNotExist:
        cajero = User.objects.first()
    with transaction.atomic():
        op = OperacionVenta.objects.create(
            mesa=mesa,
            empleado_caja=cajero,
            subtotal_base=0, total_impuestos=0, total_facturado=0,
            forma_pago='EFECTIVO',
        )
        subtotal = Decimal('0')
        for item in items:
            try:
                art = Articulo.objects.get(id=item['id'])
                cant = int(item.get('cantidad', 1))
                lv = LineaVenta.objects.create(
                    venta=op, articulo=art, unidades=cant,
                    precio_aplicado_con_iva=art.precio_con_iva,
                )
                subtotal += art.precio_con_iva * cant
            except Articulo.DoesNotExist:
                continue
        iva = subtotal * Decimal('0.16')
        op.subtotal_base = subtotal
        op.total_impuestos = iva
        op.total_facturado = subtotal + iva
        op.save()
        PedidoCocina.objects.create(mesa=mesa, venta=op, estado='PENDIENTE')
    return JsonResponse({
        'ok': True, 'venta_id': op.id, 'total': float(op.total_facturado),
        'mesa': mesa.numero,
    })


# 3. RECETAS PASO A PASO EN COCINA
@login_required
@user_passes_test(es_cocina)
def api_receta_pasos(request, articulo_id):
    """Obtiene los pasos de receta para un articulo"""
    pasos = PasoReceta.objects.filter(articulo_id=articulo_id).values(
        'orden', 'titulo', 'descripcion', 'tiempo_minutos', 'consejo'
    )
    return JsonResponse({'pasos': list(pasos)})


@login_required
@user_passes_test(es_gerente)
def api_crear_paso_receta(request):
    """Crea un paso de receta nuevo"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    articulo_id = data.get('articulo_id')
    try:
        articulo = Articulo.objects.get(id=articulo_id)
    except Articulo.DoesNotExist:
        return JsonResponse({'error': 'Articulo no encontrado'}, status=404)
    ultimo_orden = PasoReceta.objects.filter(articulo=articulo).count()
    paso = PasoReceta.objects.create(
        articulo=articulo,
        orden=ultimo_orden + 1,
        titulo=data.get('titulo', ''),
        descripcion=data.get('descripcion', ''),
        tiempo_minutos=data.get('tiempo_minutos', 0),
        consejo=data.get('consejo', ''),
    )
    return JsonResponse({'ok': True, 'id': paso.id})


# 4. CALCULADORA DE PROPINAS
@login_required
@user_passes_test(es_empleado)
def api_calcular_propina(request):
    """Calcula propina y total"""
    subtotal = Decimal(request.GET.get('subtotal', '0'))
    porcentaje = Decimal(request.GET.get('porcentaje', '15'))
    propina = subtotal * (porcentaje / Decimal('100'))
    total = subtotal + propina
    return JsonResponse({
        'subtotal': float(subtotal),
        'propina': float(propina),
        'porcentaje': float(porcentaje),
        'total': float(total),
    })


# 5. MAPA DE CALOR EN VIVO
@login_required
@user_passes_test(es_gerente)
def api_heatmap_tiempo_real(request):
    """Heatmap de ventas en tiempo real (ultima hora)"""
    ahora = timezone.now()
    hace_1h = ahora - timedelta(hours=1)
    ventas_recientes = (
        LineaVenta.objects
        .filter(venta__fecha_registro__gte=hace_1h)
        .values('articulo__nombre', 'articulo__categoria__nombre')
        .annotate(
            unidades=Sum('unidades'),
            ingresos=Sum('precio_aplicado_con_iva'),
        )
        .order_by('-unidades')[:15]
    )
    heatmap = []
    for v in ventas_recientes:
        heatmap.append({
            'producto': v['articulo__nombre'],
            'categoria': v['articulo__categoria__nombre'] or 'Sin categoría',
            'unidades': v['unidades'],
            'ingresos': float(v['ingresos'] or 0),
            'intensidad': min(100, int((v['unidades'] or 0) * 20)),
        })
    total_ventas = OperacionVenta.objects.filter(
        fecha_registro__gte=hace_1h
    ).count()
    total_ingresos = float(
        OperacionVenta.objects.filter(
            fecha_registro__gte=hace_1h
        ).aggregate(s=Sum('total_facturado'))['s'] or 0
    )
    return JsonResponse({
        'heatmap': heatmap,
        'total_ventas_hora': total_ventas,
        'total_ingresos_hora': total_ingresos,
    })


# 6. PANTALLA TV PRODUCTOS
@login_required
@user_passes_test(es_gerente)
def vista_tv_display(request):
    """Pantalla para TV del local con productos en rotacion"""
    articulos = Articulo.objects.filter(activo=True).select_related('categoria').order_by('nombre')[:20]
    return render(request, 'tv_display.html', {'articulos': list(articulos)})


# 7. NOTIFICACIONES PUSH
@login_required
@user_passes_test(es_empleado)
def api_crear_notificacion(request):
    """Crea una notificacion push"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    NotificacionPush.objects.create(
        titulo=data.get('titulo', ''),
        mensaje=data.get('mensaje', ''),
        url=data.get('url', '/'),
    )
    return JsonResponse({'ok': True})


@login_required
@user_passes_test(es_empleado)
def api_notificaciones_pendientes(request):
    """Devuelve notificaciones pendientes"""
    notifs = NotificacionPush.objects.filter(enviada=False).order_by('fecha')[:10]
    result = []
    for n in notifs:
        result.append({
            'id': n.id,
            'titulo': n.titulo,
            'mensaje': n.mensaje,
            'url': n.url,
        })
        n.enviada = True
        n.save(update_fields=['enviada'])
    return JsonResponse({'notificaciones': result})


# 8. BACKUP USB AUTOMATICO
@login_required
@user_passes_test(es_gerente)
def api_backup_usb(request):
    """Genera backup de la base de datos"""
    import os, subprocess
    import os
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(backup_dir, f'backup_{timestamp}.json')
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    result = subprocess.run(
        ['python', 'manage.py', 'dumpdata', 'tpv', '--indent', '2'],
        capture_output=True, cwd=str(settings.BASE_DIR), env=env
    )
    if result.returncode == 0:
        data = result.stdout.decode('utf-8', errors='replace')
        with open(backup_file, 'w', encoding='utf-8') as f:
            f.write(data)
        return JsonResponse({
            'ok': True,
            'archivo': backup_file,
            'tamano': os.path.getsize(backup_file),
            'timestamp': timestamp,
        })
    else:
        return JsonResponse({'error': result.stderr.decode('utf-8', errors='replace')}, status=500)


# ==================== 8 NUEVAS FEATURES (ROUND 5) ====================

# 1. DESCUENTOS INTELIGENTES
@login_required
@user_passes_test(es_empleado)
def api_descuentos_activos(request):
    """Devuelve descuentos inteligentes activos segun hora y dia"""
    ahora = timezone.localtime(timezone.now())
    hora_actual = ahora.time()
    dia_semana = ahora.weekday()
    descuentos = DescuentoInteligente.objects.filter(activo=True)
    resultado = []
    for d in descuentos:
        dias = [int(x) for x in d.dias_semana.split(',')]
        if dia_semana not in dias:
            continue
        if d.hora_inicio and d.hora_fin:
            if d.hora_inicio <= d.hora_fin:
                if not (d.hora_inicio <= hora_actual <= d.hora_fin):
                    continue
            else:
                if not (hora_actual >= d.hora_inicio or hora_actual <= d.hora_fin):
                    continue
        info = {
            'id': d.id,
            'nombre': d.nombre,
            'tipo': d.tipo,
            'porcentaje': float(d.porcentaje),
            'cantidad_minima': d.cantidad_minima,
        }
        if d.articulo:
            info['articulo'] = d.articulo.nombre
        resultado.append(info)
    return JsonResponse({'descuentos': resultado})


@login_required
@user_passes_test(es_empleado)
def api_aplicar_descuento(request):
    """Aplica descuento inteligente a items del ticket"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    items = data.get('items', [])
    descuento_total = Decimal('0')
    detalles = []
    ahora = timezone.localtime(timezone.now())
    hora_actual = ahora.time()
    for item in items:
        try:
            art = Articulo.objects.get(id=item['id'])
            cant = int(item.get('cantidad', 1))
            descuentos_art = DescuentoInteligente.objects.filter(
                activo=True, articulo=art
            )
            for d in descuentos_art:
                if d.tipo == '3X2' and cant >= 3:
                    desc = art.precio_con_iva * (cant // 3)
                    descuento_total += desc
                    detalles.append(f"3x2 en {art.nombre}: -{desc}€")
                elif d.tipo == '2DA_50' and cant >= 2:
                    desc = art.precio_con_iva * Decimal('0.5')
                    descuento_total += desc
                    detalles.append(f"2da unidad {art.nombre}: -{desc}€")
                elif d.tipo == 'HAPPY_HOUR':
                    if d.hora_inicio and d.hora_fin:
                        if d.hora_inicio <= hora_actual <= d.hora_fin:
                            desc = art.precio_con_iva * cant * (d.porcentaje / Decimal('100'))
                            descuento_total += desc
                            detalles.append(f"Happy Hour {art.nombre}: -{desc}€")
        except Articulo.DoesNotExist:
            continue
    return JsonResponse({
        'descuento_total': float(descuento_total),
        'detalles': detalles,
    })


# 2. TRACKING CUMPLEANOS
@login_required
@user_passes_test(es_empleado)
def api_cumpleanos_hoy(request):
    """Devuelve clientes que cumplen anos hoy o esta semana"""
    hoy = timezone.now().date()
    clientes_hoy = Cliente.objects.filter(
        fecha_nacimiento__month=hoy.month,
        fecha_nacimiento__day=hoy.day,
    )
    desde = hoy
    hasta = hoy + timedelta(days=7)
    clientes_semana = Cliente.objects.filter(
        fecha_nacimiento__month=hoy.month,
        fecha_nacimiento__day__gte=hoy.day,
        fecha_nacimiento__day__lte=hasta.day,
    ).exclude(id__in=clientes_hoy)
    hoy_list = []
    for c in clientes_hoy:
        hoy_list.append({
            'id': c.id, 'nombre': c.nombre, 'email': c.email,
            'puntos': c.puntos_fidelidad, 'nivel': c.nivel,
        })
    semana_list = []
    for c in clientes_semana:
        semana_list.append({
            'id': c.id, 'nombre': c.nombre,
            'dia': c.fecha_nacimiento.day,
        })
    return JsonResponse({'hoy': hoy_list, 'semana': semana_list})


# 3. RATING PRODUCTOS
@login_required
@user_passes_test(es_empleado)
def api_rating_producto(request, articulo_id):
    """Obtiene ratings de un producto"""
    ratings = RatingProducto.objects.filter(articulo_id=articulo_id)[:20]
    result = []
    for r in ratings:
        result.append({
            'nombre': r.cliente_nombre,
            'estrellas': r.estrellas,
            'comentario': r.comentario,
            'fecha': r.fecha.strftime('%d/%m %H:%M'),
        })
    promedio = ratings.aggregate(
        prom=Avg('estrellas')
    )['prom'] if ratings.exists() else 0
    return JsonResponse({'ratings': result, 'promedio': float(promedio or 0), 'total': ratings.count()})


@login_required
@user_passes_test(es_empleado)
def api_crear_rating(request):
    """Crea un rating de producto"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    try:
        art = Articulo.objects.get(id=data.get('articulo_id'))
        RatingProducto.objects.create(
            articulo=art,
            cliente_nombre=data.get('nombre', 'Anonimo'),
            estrellas=int(data.get('estrellas', 5)),
            comentario=data.get('comentario', ''),
        )
        return JsonResponse({'ok': True})
    except Articulo.DoesNotExist:
        return JsonResponse({'error': 'Articulo no encontrado'}, status=404)


@login_required
@user_passes_test(es_gerente)
def api_top_ratings(request):
    """Top productos mejor puntuados"""
    top = (
        RatingProducto.objects
        .values('articulo__nombre', 'articulo__id')
        .annotate(prom=Avg('estrellas'), total=Count('id'))
        .order_by('-prom')[:10]
    )
    result = []
    for t in top:
        result.append({
            'producto': t['articulo__nombre'],
            'id': t['articulo__id'],
            'promedio': float(t['prom'] or 0),
            'total_ratings': t['total'],
        })
    return JsonResponse({'top': result})


# 4. BUILDER DE COMBOS
@login_required
@user_passes_test(es_empleado)
def api_combos_disponibles(request):
    """Lista combos activos"""
    combos = Combo.objects.filter(activo=True)
    resultado = []
    ahora = timezone.localtime(timezone.now()).time()
    for c in combos:
        if c.hora_inicio and c.hora_fin:
            if not (c.hora_inicio <= ahora <= c.hora_fin):
                continue
        items = ComboItem.objects.filter(combo=c).values(
            'articulo__id', 'articulo__nombre', 'cantidad'
        )
        precio_original = 0
        item_list = []
        for i in items:
            art = Articulo.objects.filter(id=i['articulo__id']).first()
            if art:
                precio_original += float(art.precio_con_iva) * i['cantidad']
                item_list.append({
                    'nombre': i['articulo__nombre'],
                    'cantidad': i['cantidad'],
                    'precio_unitario': float(art.precio_con_iva),
                })
        ahorro = precio_original - float(c.precio)
        resultado.append({
            'id': c.id,
            'nombre': c.nombre,
            'descripcion': c.descripcion,
            'precio': float(c.precio),
            'precio_original': precio_original,
            'ahorro': ahorro,
            'items': item_list,
        })
    return JsonResponse({'combos': resultado})


@login_required
@user_passes_test(es_empleado)
def api_crear_combo(request):
    """Crea un combo nuevo"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    combo = Combo.objects.create(
        nombre=data.get('nombre', 'Combo'),
        descripcion=data.get('descripcion', ''),
        precio=Decimal(str(data.get('precio', 0))),
    )
    for item in data.get('items', []):
        try:
            art = Articulo.objects.get(id=item['articulo_id'])
            ComboItem.objects.create(
                combo=combo, articulo=art,
                cantidad=item.get('cantidad', 1),
            )
        except Articulo.DoesNotExist:
            continue
    return JsonResponse({'ok': True, 'combo_id': combo.id})


# 5. MAPA MESAS EN TIEMPO REAL
@login_required
@user_passes_test(es_empleado)
def api_mapa_mesas(request):
    """Estado en tiempo real de todas las mesas"""
    mesas = Mesa.objects.filter(activa=True).order_by('numero')
    resultado = []
    for m in mesas:
        ultima_venta = OperacionVenta.objects.filter(
            mesa=m
        ).order_by('-fecha_registro').first()
        total_venta = OperacionVenta.objects.filter(
            mesa=m, fecha_registro__date=timezone.now().date()
        ).aggregate(s=Sum('total_facturado'))['s'] or 0
        pedidos_pendientes = PedidoCocina.objects.filter(
            venta__mesa=m, estado__in=['PENDIENTE', 'PREPARANDO']
        ).count()
        tiempo_ocupada = 0
        if ultima_venta:
            tiempo_ocupada = (timezone.now() - ultima_venta.fecha_registro).total_seconds() / 60
        resultado.append({
            'id': m.id,
            'numero': m.numero,
            'activa': m.activa,
            'ocupada': pedidos_pendientes > 0 or (ultima_venta and tiempo_ocupada < 120),
            'pedidos_pendientes': pedidos_pendientes,
            'total_hoy': float(total_venta),
            'tiempo_minutos': round(tiempo_ocupada),
        })
    return JsonResponse({'mesas': resultado})


# 6. CALCULADORA DE CAMBIO
@login_required
@user_passes_test(es_empleado)
def api_calcular_cambio(request):
    """Calcula el cambio para un pago en efectivo"""
    total = Decimal(request.GET.get('total', '0'))
    pagado = Decimal(request.GET.get('pagado', '0'))
    cambio = pagado - total
    return JsonResponse({
        'total': float(total),
        'pagado': float(pagado),
        'cambio': float(max(cambio, 0)),
        'suficiente': cambio >= 0,
    })


# 7. REVIEWS RESTAURANTE
@login_required
def vista_reviews(request):
    """Pagina de reviews del restaurante"""
    reviews = ReviewRestaurante.objects.filter(visible=True)[:20]
    promedio = reviews.aggregate(prom=Avg('estrellas'))['prom'] or 0
    return render(request, 'reviews.html', {
        'reviews': reviews,
        'promedio': round(float(promedio), 1),
    })


@login_required
def api_reviews(request):
    """API de reviews"""
    reviews = ReviewRestaurante.objects.filter(visible=True)[:30]
    result = []
    for r in reviews:
        result.append({
            'nombre': r.nombre,
            'estrellas': r.estrellas,
            'comentario': r.comentario,
            'fecha': r.fecha.strftime('%d/%m/%Y'),
        })
    promedio = ReviewRestaurante.objects.filter(visible=True).aggregate(
        prom=Avg('estrellas')
    )['prom'] or 0
    total = ReviewRestaurante.objects.filter(visible=True).count()
    return JsonResponse({'reviews': result, 'promedio': round(float(promedio), 1), 'total': total})


@login_required
def api_crear_review(request):
    """Crea un review del restaurante"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST'}, status=405)
    data = json.loads(request.body)
    ReviewRestaurante.objects.create(
        nombre=data.get('nombre', 'Anonimo'),
        email=data.get('email', ''),
        estrellas=int(data.get('estrellas', 5)),
        comentario=data.get('comentario', ''),
    )
    return JsonResponse({'ok': True})


# 8. SYNC GOOGLE SHEETS
@login_required
@user_passes_test(es_gerente)
def api_sheets_sync(request):
    """Genera CSV listo para importar a Google Sheets"""
    import csv
    import io
    hoy = timezone.now().date()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Fecha', 'Hora', 'Cajero', 'Mesa', 'Subtotal', 'IVA', 'Total', 'Metodo Pago'])
    ventas = OperacionVenta.objects.filter(fecha_registro__date=hoy)
    for v in ventas:
        writer.writerow([
            v.fecha_registro.strftime('%d/%m/%Y'),
            v.fecha_registro.strftime('%H:%M'),
            str(v.empleado_caja),
            f"Mesa {v.mesa.numero}" if v.mesa else 'N/A',
            float(v.subtotal_base),
            float(v.total_impuestos),
            float(v.total_facturado),
            v.forma_pago,
        ])
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="tpv_sheets_{hoy.strftime("%Y%m%d")}.csv"'
    return response


# ==================== ADMIN PANEL: STOCK, INFORME SEMANAL, PREVISION ====================

@login_required
@user_passes_test(es_admin_panel)
def admin_stock_dashboard(request):
    """Dashboard de gestión de stock accesible desde el admin"""
    from decimal import Decimal

    insumos = InsumoMateriaPrima.objects.all().order_by('nombre')

    total_insumos = insumos.count()
    stock_bajo = insumos.filter(cantidad_actual__lte=F('cantidad_minima')).count()
    stock_critico = insumos.filter(cantidad_actual__lte=0).count()
    stock_ok = total_insumos - stock_bajo

    insumos_data = []
    for insumo in insumos:
        if insumo.cantidad_minima > 0:
            porcentaje = min(100, int((insumo.cantidad_actual / insumo.cantidad_minima) * 100))
        else:
            porcentaje = 100

        if insumo.cantidad_actual <= 0:
            estado = 'AGOTADO'
            estado_clase = 'danger'
        elif insumo.cantidad_actual <= insumo.cantidad_minima:
            estado = 'BAJO'
            estado_clase = 'warning'
        else:
            estado = 'OK'
            estado_clase = 'success'

        insumos_data.append({
            'insumo': insumo,
            'porcentaje': porcentaje,
            'estado': estado,
            'estado_clase': estado_clase,
        })

    pedidos_pendientes = PedidoProveedor.objects.filter(estado='PENDIENTE').count()

    context = {
        **admin.site.each_context(request),
        'titulo': 'Dashboard de Stock',
        'insumos_data': insumos_data,
        'total_insumos': total_insumos,
        'stock_bajo': stock_bajo,
        'stock_critico': stock_critico,
        'stock_ok': stock_ok,
        'pedidos_pendientes': pedidos_pendientes,
        'app_list': admin.site.get_app_list(request),
    }
    return render(request, 'admin/stock_dashboard.html', context)


@login_required
@user_passes_test(es_admin_panel)
def admin_informe_semanal(request):
    """Informe semanal de gastos y ventas accesible desde el admin"""
    from decimal import Decimal

    hoy = timezone.now().date()
    semana_offset = int(request.GET.get('semana', 0))
    fecha_inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
    fecha_fin_semana = fecha_inicio_semana + timedelta(days=6)

    ventas_semana = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio_semana,
        fecha_registro__date__lte=fecha_fin_semana
    )
    gastos_semana = RegistroGasto.objects.filter(
        fecha_gasto__date__gte=fecha_inicio_semana,
        fecha_gasto__date__lte=fecha_fin_semana
    )

    total_ventas = ventas_semana.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    total_gastos = gastos_semana.aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
    balance = total_ventas - total_gastos
    total_tickets = ventas_semana.count()
    ticket_medio = (total_ventas / total_tickets) if total_tickets > 0 else Decimal('0.00')

    efectivo = ventas_semana.filter(forma_pago='EFECTIVO').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tarjeta = ventas_semana.filter(forma_pago='TARJETA').aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')

    # Detalle diario
    dias_semana = []
    nombres_dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    for i in range(7):
        dia = fecha_inicio_semana + timedelta(days=i)
        ventas_dia = ventas_semana.filter(fecha_registro__date=dia)
        gastos_dia = gastos_semana.filter(fecha_gasto__date=dia)
        total_dia = ventas_dia.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        gasto_dia = gastos_dia.aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
        dias_semana.append({
            'nombre': nombres_dias[i],
            'fecha': dia,
            'ventas': total_dia,
            'gastos': gasto_dia,
            'balance': total_dia - gasto_dia,
            'tickets': ventas_dia.count(),
        })

    # Top productos vendidos
    top_productos = (
        LineaVenta.objects
        .filter(venta__fecha_registro__date__gte=fecha_inicio_semana,
                venta__fecha_registro__date__lte=fecha_fin_semana)
        .values('articulo__nombre')
        .annotate(unidades=Sum('unidades'), ingresos=Sum('precio_aplicado_con_iva'))
        .order_by('-ingresos')[:10]
    )

    # Desglose de gastos
    desglose_gastos = (
        gastos_semana
        .values('concepto')
        .annotate(total=Sum('importe_total'))
        .order_by('-total')
    )

    # Semana anterior para comparativa
    fecha_inicio_ant = fecha_inicio_semana - timedelta(days=7)
    fecha_fin_ant = fecha_fin_semana - timedelta(days=7)
    ventas_anterior = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio_ant,
        fecha_registro__date__lte=fecha_fin_ant
    )
    total_ventas_anterior = ventas_anterior.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tickets_anterior = ventas_anterior.count()
    variacion_ventas = 0
    if total_ventas_anterior > 0:
        variacion_ventas = round(((float(total_ventas) - float(total_ventas_anterior)) / float(total_ventas_anterior)) * 100, 1)

    context = {
        **admin.site.each_context(request),
        'titulo': 'Informe Semanal',
        'fecha_inicio': fecha_inicio_semana,
        'fecha_fin': fecha_fin_semana,
        'semana_offset': semana_offset,
        'total_ventas': total_ventas,
        'total_gastos': total_gastos,
        'balance': balance,
        'total_tickets': total_tickets,
        'ticket_medio': ticket_medio,
        'efectivo': efectivo,
        'tarjeta': tarjeta,
        'dias_semana': dias_semana,
        'top_productos': top_productos,
        'desglose_gastos': desglose_gastos,
        'total_ventas_anterior': total_ventas_anterior,
        'tickets_anterior': tickets_anterior,
        'variacion_ventas': variacion_ventas,
        'app_list': admin.site.get_app_list(request),
    }
    return render(request, 'admin/informe_semanal.html', context)


@login_required
@user_passes_test(es_admin_panel)
def admin_prevision_compras(request):
    """Previsión de compras del lunes basada en la semana anterior"""
    from decimal import Decimal

    hoy = timezone.now().date()
    semana_offset = int(request.GET.get('semana', 0))

    # Semana a analizar (la anterior)
    fecha_inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset - 1)
    fecha_fin_semana = fecha_inicio_semana + timedelta(days=6)

    # Semana de previsión (la siguiente)
    fecha_inicio_prevision = fecha_fin_semana + timedelta(days=1)
    fecha_fin_prevision = fecha_inicio_prevision + timedelta(days=6)

    # Ventas de la semana analizada
    ventas_semana = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio_semana,
        fecha_registro__date__lte=fecha_fin_semana
    )

    # Calcular consumo de cada insumo basado en ventas × recetas
    lineas_venta = LineaVenta.objects.filter(
        venta__in=ventas_semana
    ).select_related('articulo')

    consumo_insumos = {}
    for linea in lineas_venta:
        recetas = ComposicionReceta.objects.filter(articulo=linea.articulo)
        for receta in recetas:
            insumo_id = receta.insumo.id
            consumo = receta.cantidad_consumida * linea.unidades
            if insumo_id in consumo_insumos:
                consumo_insumos[insumo_id]['consumo_total'] += consumo
            else:
                consumo_insumos[insumo_id] = {
                    'insumo': receta.insumo,
                    'consumo_total': consumo,
                }

    # Calcular previsión para la próxima semana
    dias_analizados = 7
    dias_prevision = 7

    prevision_data = []
    total_compra = Decimal('0.00')

    for insumo in InsumoMateriaPrima.objects.all().order_by('nombre'):
        consumo_info = consumo_insumos.get(insumo.id, {'consumo_total': Decimal('0.00')})
        consumo_semanal = consumo_info['consumo_total']

        # Consumo diario promedio
        consumo_diario = consumo_semanal / dias_analizados if dias_analizados > 0 else Decimal('0.00')

        # Consumo estimado para la próxima semana
        consumo_estimado = consumo_diario * dias_prevision

        # Stock actual
        stock_actual = insumo.cantidad_actual

        # Stock proyectado al final de la próxima semana
        stock_proyectado = stock_actual - consumo_estimado

        # Cantidad a recomendar para comprar
        if stock_proyectado < insumo.cantidad_minima:
            cantidad_comprar = insumo.cantidad_minima * 2 - stock_proyectado
            if cantidad_comprar < 0:
                cantidad_comprar = Decimal('0.00')
        else:
            cantidad_comprar = Decimal('0.00')

        total_compra += cantidad_comprar

        prevision_data.append({
            'insumo': insumo,
            'consumo_semanal': consumo_semanal,
            'consumo_diario': consumo_diario,
            'consumo_estimado': consumo_estimado,
            'stock_actual': stock_actual,
            'stock_proyectado': stock_proyectado,
            'cantidad_comprar': cantidad_comprar,
            'necesita_compra': cantidad_comprar > 0,
        })

    context = {
        **admin.site.each_context(request),
        'titulo': 'Previsión de Compras',
        'semana_offset': semana_offset,
        'fecha_inicio_analisis': fecha_inicio_semana,
        'fecha_fin_analisis': fecha_fin_semana,
        'fecha_inicio_prevision': fecha_inicio_prevision,
        'fecha_fin_prevision': fecha_fin_prevision,
        'prevision_data': prevision_data,
        'total_compra': total_compra,
        'total_ventas_semana': ventas_semana.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00'),
        'app_list': admin.site.get_app_list(request),
    }
    return render(request, 'admin/prevision_compras.html', context)


# ==================== EXPORTACIONES EXCEL/PDF PARA ADMIN ====================

@login_required
@user_passes_test(es_admin_panel)
def admin_exportar_stock_excel(request):
    """Exporta el estado del stock actual a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Hoja 1: Stock Actual
    ws = wb.active
    ws.title = "Stock Actual"
    headers = ['Insumo', 'Stock Actual', 'Stock Mínimo', 'Unidad', 'Estado', 'Necesita Reposición']
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    insumos = InsumoMateriaPrima.objects.all().order_by('nombre')
    for row, insumo in enumerate(insumos, 2):
        estado = 'AGOTADO' if insumo.cantidad_actual <= 0 else ('BAJO' if insumo.necesita_reposicion else 'OK')
        ws.cell(row=row, column=1, value=insumo.nombre).border = thin_border
        ws.cell(row=row, column=2, value=float(insumo.cantidad_actual)).border = thin_border
        ws.cell(row=row, column=3, value=float(insumo.cantidad_minima)).border = thin_border
        ws.cell(row=row, column=4, value=insumo.get_unidad_medida_display()).border = thin_border
        c_estado = ws.cell(row=row, column=5, value=estado)
        c_estado.border = thin_border
        if estado == 'AGOTADO':
            c_estado.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            c_estado.font = Font(color="FFFFFF", bold=True)
        elif estado == 'BAJO':
            c_estado.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            c_estado.font = Font(bold=True)
        else:
            c_estado.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        ws.cell(row=row, column=6, value='SÍ' if insumo.necesita_reposicion else 'NO').border = thin_border

    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="stock_actual_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin_panel)
def admin_exportar_informe_semanal_excel(request):
    """Exporta el informe semanal a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from decimal import Decimal

    hoy = timezone.now().date()
    semana_offset = int(request.GET.get('semana', 0))
    fecha_inicio = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
    fecha_fin = fecha_inicio + timedelta(days=6)

    wb = openpyxl.Workbook()

    # Hoja 1: Resumen Semanal
    ws = wb.active
    ws.title = "Resumen Semanal"
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    resumen_data = [
        ['Concepto', 'Valor'],
        ['Semana del', fecha_inicio.strftime('%d/%m/%Y')],
        ['al', fecha_fin.strftime('%d/%m/%Y')],
        ['', ''],
    ]
    for row, data in enumerate(resumen_data, 1):
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            if row == 1:
                c.fill = header_fill
                c.font = header_font

    ventas = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio, fecha_registro__date__lte=fecha_fin
    )
    gastos = RegistroGasto.objects.filter(
        fecha_gasto__date__gte=fecha_inicio, fecha_gasto__date__lte=fecha_fin
    )
    total_v = ventas.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    total_g = gastos.aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')

    metrics = [
        ['Total Facturado', f'{total_v:.2f} EUR'],
        ['Total Gastos', f'{total_g:.2f} EUR'],
        ['Balance', f'{(total_v - total_g):.2f} EUR'],
        ['Tickets', ventas.count()],
        ['Ticket Medio', f'{(total_v / ventas.count() if ventas.count() > 0 else 0):.2f} EUR'],
    ]
    for row, data in enumerate(metrics, 5):
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25

    # Hoja 2: Detalle Diario
    ws2 = wb.create_sheet("Detalle Diario")
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    headers2 = ['Día', 'Fecha', 'Ventas', 'Gastos', 'Balance', 'Tickets']
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
        c.font = header_font

    for i in range(7):
        dia = fecha_inicio + timedelta(days=i)
        v_dia = ventas.filter(fecha_registro__date=dia).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        g_dia = gastos.filter(fecha_gasto__date=dia).aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
        row = i + 2
        ws2.cell(row=row, column=1, value=dias[i])
        ws2.cell(row=row, column=2, value=dia.strftime('%d/%m'))
        ws2.cell(row=row, column=3, value=float(v_dia))
        ws2.cell(row=row, column=4, value=float(g_dia))
        ws2.cell(row=row, column=5, value=float(v_dia - g_dia))
        ws2.cell(row=row, column=6, value=ventas.filter(fecha_registro__date=dia).count())

    for col in range(1, 7):
        ws2.column_dimensions[chr(64+col)].width = 18

    # Hoja 3: Top Productos
    ws3 = wb.create_sheet("Top Productos")
    headers3 = ['Producto', 'Unidades Vendidas', 'Ingresos']
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        c.font = header_font

    top = (
        LineaVenta.objects
        .filter(venta__fecha_registro__date__gte=fecha_inicio, venta__fecha_registro__date__lte=fecha_fin)
        .values('articulo__nombre')
        .annotate(unidades=Sum('unidades'), ingresos=Sum('precio_aplicado_con_iva'))
        .order_by('-ingresos')[:10]
    )
    for row, t in enumerate(top, 2):
        ws3.cell(row=row, column=1, value=t['articulo__nombre'])
        ws3.cell(row=row, column=2, value=t['unidades'])
        ws3.cell(row=row, column=3, value=float(t['ingresos'] or 0))

    for col in range(1, 4):
        ws3.column_dimensions[chr(64+col)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="informe_semanal_{fecha_inicio.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin_panel)
def admin_exportar_prevision_excel(request):
    """Exporta la previsión de compras a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from decimal import Decimal

    hoy = timezone.now().date()
    semana_offset = int(request.GET.get('semana', 0))

    fecha_inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset - 1)
    fecha_fin_semana = fecha_inicio_semana + timedelta(days=6)
    fecha_inicio_prevision = fecha_fin_semana + timedelta(days=1)
    fecha_fin_prevision = fecha_inicio_prevision + timedelta(days=6)

    ventas_semana = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio_semana,
        fecha_registro__date__lte=fecha_fin_semana
    )

    lineas_venta = LineaVenta.objects.filter(venta__in=ventas_semana).select_related('articulo')

    consumo_insumos = {}
    for linea in lineas_venta:
        recetas = ComposicionReceta.objects.filter(articulo=linea.articulo)
        for receta in recetas:
            insumo_id = receta.insumo.id
            consumo = receta.cantidad_consumida * linea.unidades
            if insumo_id in consumo_insumos:
                consumo_insumos[insumo_id] += consumo
            else:
                consumo_insumos[insumo_id] = consumo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"

    headers = ['Insumo', 'Unidad', 'Stock Actual', 'Consumo Semanal', 'Stock Proyectado', 'Cantidad a Comprar']
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    row = 2
    total_compra = Decimal('0.00')
    for insumo in InsumoMateriaPrima.objects.all().order_by('nombre'):
        consumo = consumo_insumos.get(insumo.id, Decimal('0.00'))
        consumo_diario = consumo / 7
        consumo_estimado = consumo_diario * 7
        stock_proyectado = insumo.cantidad_actual - consumo_estimado

        if stock_proyectado < insumo.cantidad_minima:
            cantidad_comprar = insumo.cantidad_minima * 2 - stock_proyectado
            if cantidad_comprar < 0:
                cantidad_comprar = Decimal('0.00')
        else:
            cantidad_comprar = Decimal('0.00')

        total_compra += cantidad_comprar

        ws.cell(row=row, column=1, value=insumo.nombre).border = thin_border
        ws.cell(row=row, column=2, value=insumo.unidad_medida).border = thin_border
        ws.cell(row=row, column=3, value=float(insumo.cantidad_actual)).border = thin_border
        ws.cell(row=row, column=4, value=float(consumo)).border = thin_border
        ws.cell(row=row, column=5, value=float(stock_proyectado)).border = thin_border
        c_comprar = ws.cell(row=row, column=6, value=float(cantidad_comprar))
        c_comprar.border = thin_border
        if cantidad_comprar > 0:
            c_comprar.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            c_comprar.font = Font(bold=True)
        row += 1

    # Fila de total
    ws.cell(row=row + 1, column=5, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=row + 1, column=6, value=float(total_compra)).font = Font(bold=True, color="FF0000")

    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="prevision_compras_{fecha_inicio_semana.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin_panel)
def admin_exportar_informe_semanal_pdf(request):
    """Exporta el informe semanal a PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    import io
    from decimal import Decimal

    hoy = timezone.now().date()
    semana_offset = int(request.GET.get('semana', 0))
    fecha_inicio = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
    fecha_fin = fecha_inicio + timedelta(days=6)

    ventas = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio, fecha_registro__date__lte=fecha_fin
    )
    gastos = RegistroGasto.objects.filter(
        fecha_gasto__date__gte=fecha_inicio, fecha_gasto__date__lte=fecha_fin
    )
    total_v = ventas.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    total_g = gastos.aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    y = h - 40

    # Título
    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(w/2, y, 'INFORME SEMANAL - TPV Cafeteria')
    y -= 25
    c.setFont('Helvetica', 11)
    c.drawCentredString(w/2, y, f'Semana: {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}')
    y -= 35

    # Resumen
    c.setFont('Helvetica-Bold', 14)
    c.drawString(40, y, 'RESUMEN SEMANAL')
    y -= 25
    c.setFont('Helvetica', 11)
    for label, val in [
        ('Total Facturado', f'{total_v:.2f} EUR'),
        ('Total Gastos', f'{total_g:.2f} EUR'),
        ('Balance Neto', f'{(total_v - total_g):.2f} EUR'),
        ('Tickets Totales', str(ventas.count())),
        ('Ticket Medio', f'{(total_v / ventas.count() if ventas.count() > 0 else 0):.2f} EUR'),
    ]:
        c.drawString(50, y, f'{label}:')
        c.drawRightString(w - 40, y, val)
        y -= 18

    # Detalle diario
    c.setFont('Helvetica-Bold', 14)
    c.drawString(40, y, 'DETALLE POR DIA')
    y -= 25
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, y, 'Dia')
    c.drawString(100, y, 'Fecha')
    c.drawString(160, y, 'Ventas')
    c.drawRightString(w - 40, y, 'Gastos')
    y -= 15
    c.line(40, y, w - 40, y)
    y -= 15

    dias = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado', 'Domingo']
    c.setFont('Helvetica', 10)
    for i in range(7):
        dia = fecha_inicio + timedelta(days=i)
        v_dia = ventas.filter(fecha_registro__date=dia).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
        g_dia = gastos.filter(fecha_gasto__date=dia).aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
        c.drawString(40, y, dias[i])
        c.drawString(100, y, dia.strftime('%d/%m'))
        c.drawString(160, y, f'{v_dia:.2f}')
        c.drawRightString(w - 40, y, f'{g_dia:.2f}')
        y -= 15

    # Top productos
    y -= 15
    c.setFont('Helvetica-Bold', 14)
    c.drawString(40, y, 'TOP 10 PRODUCTOS')
    y -= 25
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, y, '#')
    c.drawString(60, y, 'Producto')
    c.drawRightString(w - 40, y, 'Ingresos')
    y -= 15
    c.line(40, y, w - 40, y)
    y -= 15

    top = (
        LineaVenta.objects
        .filter(venta__fecha_registro__date__gte=fecha_inicio, venta__fecha_registro__date__lte=fecha_fin)
        .values('articulo__nombre')
        .annotate(unidades=Sum('unidades'), ingresos=Sum('precio_aplicado_con_iva'))
        .order_by('-ingresos')[:10]
    )
    c.setFont('Helvetica', 10)
    for idx, t in enumerate(top, 1):
        if y < 50:
            c.showPage()
            y = h - 40
        c.drawString(40, y, str(idx))
        c.drawString(60, y, t['articulo__nombre'][:30])
        c.drawRightString(w - 40, y, f'{float(t["ingresos"] or 0):.2f}')
        y -= 15

    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="informe_semanal_{fecha_inicio.strftime("%Y%m%d")}.pdf"'
    return response


@login_required
@user_passes_test(es_admin_panel)
def admin_exportar_prevision_pdf(request):
    """Exporta la previsión de compras a PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    import io
    from decimal import Decimal

    hoy = timezone.now().date()
    semana_offset = int(request.GET.get('semana', 0))

    fecha_inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset - 1)
    fecha_fin_semana = fecha_inicio_semana + timedelta(days=6)
    fecha_inicio_prevision = fecha_fin_semana + timedelta(days=1)
    fecha_fin_prevision = fecha_inicio_prevision + timedelta(days=6)

    ventas_semana = OperacionVenta.objects.filter(
        fecha_registro__date__gte=fecha_inicio_semana,
        fecha_registro__date__lte=fecha_fin_semana
    )
    lineas_venta = LineaVenta.objects.filter(venta__in=ventas_semana).select_related('articulo')

    consumo_insumos = {}
    for linea in lineas_venta:
        recetas = ComposicionReceta.objects.filter(articulo=linea.articulo)
        for receta in recetas:
            insumo_id = receta.insumo.id
            consumo = receta.cantidad_consumida * linea.unidades
            if insumo_id in consumo_insumos:
                consumo_insumos[insumo_id] += consumo
            else:
                consumo_insumos[insumo_id] = consumo

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    y = h - 40

    # Título
    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(w/2, y, 'PREVISION DE COMPRAS')
    y -= 25
    c.setFont('Helvetica', 11)
    c.drawCentredString(w/2, y, f'Semana de analisis: {fecha_inicio_semana.strftime("%d/%m/%Y")} - {fecha_fin_semana.strftime("%d/%m/%Y")}')
    y -= 15
    c.drawCentredString(w/2, y, f'Prevision para: {fecha_inicio_prevision.strftime("%d/%m/%Y")} - {fecha_fin_prevision.strftime("%d/%m/%Y")}')
    y -= 35

    # Lista de compras
    c.setFont('Helvetica-Bold', 14)
    c.drawString(40, y, 'LISTA DE COMPRAS')
    y -= 25
    c.setFont('Helvetica-Bold', 9)
    c.drawString(40, y, 'Insumo')
    c.drawString(180, y, 'Unidad')
    c.drawString(240, y, 'Stock')
    c.drawString(300, y, 'Consumo')
    c.drawRightString(w - 40, y, 'Comprar')
    y -= 15
    c.line(40, y, w - 40, y)
    y -= 15

    c.setFont('Helvetica', 9)
    total_compra = Decimal('0.00')
    for insumo in InsumoMateriaPrima.objects.all().order_by('nombre'):
        if y < 50:
            c.showPage()
            y = h - 40

        consumo = consumo_insumos.get(insumo.id, Decimal('0.00'))
        consumo_diario = consumo / 7
        consumo_estimado = consumo_diario * 7
        stock_proyectado = insumo.cantidad_actual - consumo_estimado

        if stock_proyectado < insumo.cantidad_minima:
            cantidad_comprar = insumo.cantidad_minima * 2 - stock_proyectado
            if cantidad_comprar < 0:
                cantidad_comprar = Decimal('0.00')
        else:
            cantidad_comprar = Decimal('0.00')

        total_compra += cantidad_comprar

        c.drawString(40, y, insumo.nombre[:25])
        c.drawString(180, y, insumo.unidad_medida)
        c.drawString(240, y, f'{insumo.cantidad_actual:.1f}')
        c.drawString(300, y, f'{consumo:.1f}')
        c.drawRightString(w - 40, y, f'{cantidad_comprar:.1f}' if cantidad_comprar > 0 else '-')
        y -= 15

    # Total
    y -= 10
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, f'TOTAL ESTIMADO A COMPRAR: {total_compra:.2f}')

    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="prevision_compras_{fecha_inicio_semana.strftime("%Y%m%d")}.pdf"'
    return response


# ==================== DASHBOARD DE GERENCIA ====================

@login_required
@user_passes_test(es_admin_panel)
def admin_gerencia_dashboard(request):
    """Dashboard principal de gerencia con KPIs, gráficos y estadísticas"""
    import json
    from decimal import Decimal
    from django.db.models.functions import TruncHour

    hoy = timezone.now().date()
    mes_actual = hoy.month
    anyo_actual = hoy.year

    # ===== KPIs PRINCIPALES =====
    # Hoy
    ventas_hoy = OperacionVenta.objects.filter(fecha_registro__date=hoy)
    ingresos_hoy = ventas_hoy.aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tickets_hoy = ventas_hoy.count()
    ticket_medio_hoy = (ingresos_hoy / tickets_hoy) if tickets_hoy > 0 else Decimal('0.00')

    # Ayer
    ayer = hoy - timedelta(days=1)
    ingresos_ayer = OperacionVenta.objects.filter(fecha_registro__date=ayer).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tickets_ayer = OperacionVenta.objects.filter(fecha_registro__date=ayer).count()

    # Semana (lunes a hoy)
    lunes = hoy - timedelta(days=hoy.weekday())
    ingresos_semana = OperacionVenta.objects.filter(
        fecha_registro__date__gte=lunes, fecha_registro__date__lte=hoy
    ).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tickets_semana = OperacionVenta.objects.filter(
        fecha_registro__date__gte=lunes, fecha_registro__date__lte=hoy
    ).count()

    # Mes
    ingresos_mes = OperacionVenta.objects.filter(
        fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual
    ).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tickets_mes = OperacionVenta.objects.filter(
        fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual
    ).count()

    # Mes anterior (para comparar)
    if mes_actual == 1:
        mes_ant = 12
        anyo_ant = anyo_actual - 1
    else:
        mes_ant = mes_actual - 1
        anyo_ant = anyo_actual
    ingresos_mes_anterior = OperacionVenta.objects.filter(
        fecha_registro__month=mes_ant, fecha_registro__year=anyo_ant
    ).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')

    variacion_mes = 0
    if ingresos_mes_anterior > 0:
        variacion_mes = round(((float(ingresos_mes) - float(ingresos_mes_anterior)) / float(ingresos_mes_anterior)) * 100, 1)

    # ===== GRÁFICO: Ventas últimos 7 días =====
    ventas_7dias = []
    nombres_dias = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']
    for i in range(7):
        dia = hoy - timedelta(days=6-i)
        total = OperacionVenta.objects.filter(fecha_registro__date=dia).aggregate(t=Sum('total_facturado'))['t'] or 0
        ventas_7dias.append({
            'dia': nombres_dias[dia.weekday()],
            'fecha': dia.strftime('%d/%m'),
            'total': float(total)
        })

    # ===== GRÁFICO: Ventas por hora (hoy) =====
    ventas_por_hora = []
    for h in range(8, 22):  # 8am a 9pm
        total = ventas_hoy.filter(
            fecha_registro__hour=h
        ).aggregate(t=Sum('total_facturado'))['t'] or 0
        ventas_por_hora.append({'hora': f'{h}:00', 'total': float(total)})

    # ===== TOP 10 PRODUCTOS (mes) =====
    top_productos_mes = (
        LineaVenta.objects
        .filter(venta__fecha_registro__month=mes_actual, venta__fecha_registro__year=anyo_actual)
        .values('articulo__nombre')
        .annotate(unidades=Sum('unidades'), ingresos=Sum('precio_aplicado_con_iva'))
        .order_by('-ingresos')[:10]
    )

    # ===== TOP 10 PRODUCTOS (hoy) =====
    top_productos_hoy = (
        LineaVenta.objects
        .filter(venta__fecha_registro__date=hoy)
        .values('articulo__nombre')
        .annotate(unidades=Sum('unidades'), ingresos=Sum('precio_aplicado_con_iva'))
        .order_by('-ingresos')[:10]
    )

    # ===== RANKING EMPLEADOS (mes) =====
    ranking_empleados = (
        OperacionVenta.objects
        .filter(fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual)
        .values('empleado_caja__username')
        .annotate(
            total_ventas=Sum('total_facturado'),
            num_tickets=Count('id')
        )
        .order_by('-total_ventas')[:10]
    )

    # ===== MÉTODOS DE PAGO =====
    efectivo_mes = OperacionVenta.objects.filter(
        fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual, forma_pago='EFECTIVO'
    ).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')
    tarjeta_mes = OperacionVenta.objects.filter(
        fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual, forma_pago='TARJETA'
    ).aggregate(t=Sum('total_facturado'))['t'] or Decimal('0.00')

    # ===== GASTOS MES =====
    gastos_mes = RegistroGasto.objects.filter(
        fecha_gasto__month=mes_actual, fecha_gasto__year=anyo_actual
    ).aggregate(t=Sum('importe_total'))['t'] or Decimal('0.00')
    balance_mes = ingresos_mes - gastos_mes

    # ===== STOCK BAJO =====
    stock_bajo = InsumoMateriaPrima.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()

    # ===== SATISFACCIÓN =====
    from django.db.models import CharField
    from django.db.models.functions import Substr
    satisfaccion_stats = OperacionVenta.objects.filter(
        fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual,
        satisfaccion__gt=''
    ).values('satisfaccion').annotate(total=Count('id')).order_by('-total')

    # ===== MESA MÁS POPULAR =====
    mesa_popular = (
        OperacionVenta.objects
        .filter(fecha_registro__month=mes_actual, fecha_registro__year=anyo_actual, mesa__isnull=False)
        .values('mesa__numero')
        .annotate(ventas=Count('id'))
        .order_by('-ventas')[:5]
    )

    context = {
        **admin.site.each_context(request),
        'titulo': 'Dashboard de Gerencia',
        'hoy': hoy,

        # KPIs
        'ingresos_hoy': ingresos_hoy,
        'tickets_hoy': tickets_hoy,
        'ticket_medio_hoy': ticket_medio_hoy,
        'ingresos_ayer': ingresos_ayer,
        'tickets_ayer': tickets_ayer,
        'ingresos_semana': ingresos_semana,
        'tickets_semana': tickets_semana,
        'ingresos_mes': ingresos_mes,
        'tickets_mes': tickets_mes,
        'variacion_mes': variacion_mes,

        # Gráficos (JSON)
        'ventas_7dias_json': json.dumps(ventas_7dias),
        'ventas_por_hora_json': json.dumps(ventas_por_hora),

        # Tablas
        'top_productos_mes': top_productos_mes,
        'top_productos_hoy': top_productos_hoy,
        'ranking_empleados': ranking_empleados,

        # Pagos
        'efectivo_mes': efectivo_mes,
        'tarjeta_mes': tarjeta_mes,

        # Gastos
        'gastos_mes': gastos_mes,
        'balance_mes': balance_mes,

        # Stock
        'stock_bajo': stock_bajo,

        # Satisfacción
        'satisfaccion_stats': satisfaccion_stats,

        # Mesas
        'mesa_popular': mesa_popular,

        'app_list': admin.site.get_app_list(request),
    }
    return render(request, 'admin/gerencia_dashboard.html', context)
